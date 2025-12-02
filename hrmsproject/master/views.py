from datetime import datetime, date, timedelta
import json
import uuid

from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.core.validators import validate_email
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.views.decorators.http import require_POST, require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    AdditionDeduction,
    AssetType,
    Company,
    Degree,
    Department,
    Designation,
    Employee,
    EmployeeAccountInfo,
    EmployeeAssetAssignment,
    EmployeeDependent,
    EmployeeExperience,
    EmployeeQualification,
    EmployeeVehicleDetail,
    ExpenseType,
    Holiday,
    LeaveType,
    Plant,
    Shift,
    SalaryType,
    Site,
    SubExpense,
)


ROSTER_SITES = [
    {'value': 'atmakur', 'label': 'Atmakur'},
    {'value': 'punganur', 'label': 'Punganur'},
    {'value': 'kavali', 'label': 'Kavali'},
    {'value': 'perungudi', 'label': 'Perungudi'},
]

ROSTER_SALARY_TYPES = [
    {'value': 'SALARY', 'label': 'Salary'},
    {'value': 'WAGES', 'label': 'Wages'},
    {'value': 'OTHERS', 'label': 'Others'},
]

ROSTER_EMPLOYEES = [
    {
        'id': 1,
        'name': 'Kusal Bose',
        'designation': 'Earthmover Incharge',
        'default_site': 'atmakur',
    },
    {
        'id': 2,
        'name': 'Shaik Firoz',
        'designation': 'Admin Executive',
        'default_site': 'atmakur',
    },
]

WEEK_START = date(2025, 9, 14)
ROSTER_WEEK_DAYS = [
    {
        'date': WEEK_START + timedelta(days=index),
        'label': (WEEK_START + timedelta(days=index)).strftime('%A'),
    }
    for index in range(7)
]

WEEK_ROSTERS = [
    {
        'id': 1,
        'division_name': 'Punganur',
        'shift_type': 'Week Wise',
        'entry_date': date(2025, 9, 20),
        'site': 'punganur',
        'site_label': 'Punganur',
        'salary_type': 'WAGES',
        'status': 'Published',
    },
    {
        'id': 2,
        'division_name': 'Atmakur',
        'shift_type': 'Week Wise',
        'entry_date': date(2025, 9, 13),
        'site': 'atmakur',
        'site_label': 'Atmakur',
        'salary_type': 'SALARY',
        'status': 'Draft',
    },
]

HOLIDAY_SITE_OPTIONS = [
    'Poonamallee',
    'Gudalur',
    'Nellore',
    'Degenration',
    'Vijayawada',
    'Erode (Head Office)',
]

HOLIDAY_TYPE_OPTIONS = [
    'Ascent',
    'Contract',
    'Public',
    'Optional',
]

MONTH_ROSTERS = [
    {
        'id': 101,
        'division_name': 'Punganur',
        'entry_month': date(2025, 9, 1),
        'site': 'punganur',
        'site_label': 'Punganur',
        'salary_type': 'WAGES',
        'status': 'Published',
    },
    {
        'id': 102,
        'division_name': 'Kavali',
        'entry_month': date(2025, 8, 1),
        'site': 'kavali',
        'site_label': 'Kavali',
        'salary_type': 'WAGES',
        'status': 'Draft',
    },
    {
        'id': 103,
        'division_name': 'Anantapur',
        'entry_month': date(2025, 9, 1),
        'site': 'atmakur',
        'site_label': 'Atmakur',
        'salary_type': 'SALARY',
        'status': 'Published',
    },
]


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


@permission_required('master.view_company', raise_exception=True)
def company_list(request):
    companies = Company.objects.all()
    query = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()

    if query:
        companies = companies.filter(billing_name__icontains=query)
    if status:
        companies = companies.filter(status=status)

    context = {
        'companies': companies,
        'query': query,
        'status_filter': status,
    }
    return render(request, 'master/company_creation/list.html', context)


@permission_required('master.add_company', raise_exception=True)
def company_create1(request):
    if request.method == 'POST':
        submitted = request.POST.dict()
        company_group = submitted.get('company_group', '').strip()
        address = submitted.get('address', '').strip()
        billing_name = submitted.get('billing_name', '').strip()
        billing_address = submitted.get('billing_address', '').strip()
        mobile_no = submitted.get('mobile_no', '').strip()
        phone_no = submitted.get('phone_no', '').strip()
        gstin_no = submitted.get('gstin_no', '').strip()
        status = submitted.get('status', Company.STATUS_ACTIVE)
        logo = request.FILES.get('logo')

        errors = {}
        for field, value in (
            ('company_group', company_group),
            ('address', address),
            ('billing_name', billing_name),
            ('billing_address', billing_address),
            ('mobile_no', mobile_no),
            ('gstin_no', gstin_no),
        ):
            if not value:
                errors[field] = 'This field is required.'

        if gstin_no and Company.objects.filter(gstin_no=gstin_no).exists():
            errors['gstin_no'] = 'GSTIN already exists.'

        if mobile_no:
            if not mobile_no.isdigit():
                errors['mobile_no'] = 'Mobile number should contain digits only.'
            elif len(mobile_no) != 10:
                errors['mobile_no'] = 'Mobile number must be exactly 10 digits.'

        if phone_no:
            if not phone_no.isdigit():
                errors['phone_no'] = 'Phone number should contain digits only.'
            elif len(phone_no) != 10:
                errors['phone_no'] = 'Phone number must be exactly 10 digits.'

        if gstin_no:
            if len(gstin_no) != 15:
                errors['gstin_no'] = 'GSTIN number must be exactly 15 characters.'
            elif not gstin_no.isalnum():
                errors['gstin_no'] = 'GSTIN number must contain only alphanumeric characters.'

        if not errors:
            Company.objects.create(
                company_group=company_group,
                address=address,
                billing_name=billing_name,
                billing_address=billing_address,
                mobile_no=mobile_no,
                phone_no=phone_no,
                gstin_no=gstin_no,
                status=status,
                logo=logo,
            )
            messages.success(request, 'Company created successfully.')
            return redirect('master:company_list')

        return render(
            request,
            'master/company_creation/create1.html',
            {
                'errors': errors,
                'submitted': submitted,
            },
        )

    return render(request, 'master/company_creation/create1.html', {
        'errors': {},
        'submitted': {},
    })


@permission_required('master.add_company', raise_exception=True)
def company_add(request):
    if request.method == 'POST':
        submitted = request.POST.dict()
        errors = {}

        company_group = submitted.get('company_group', '').strip()
        billing_name = submitted.get('billing_name', '').strip()
        mobile_no = submitted.get('mobile_no', '').strip()
        phone_no = submitted.get('phone_no', '').strip()
        address = submitted.get('address', '').strip()
        billing_address = submitted.get('billing_address', '').strip()
        gstin_no = submitted.get('gstin_no', '').strip()
        status = submitted.get('status', Company.STATUS_ACTIVE)
        logo = request.FILES.get('logo')

        required_fields = [
            ('company_group', company_group),
            ('billing_name', billing_name),
            ('address', address),
            ('billing_address', billing_address),
            ('mobile_no', mobile_no),
            ('gstin_no', gstin_no),
        ]
        for field, value in required_fields:
            if not value:
                errors[field] = 'This field is required.'

        if gstin_no and Company.objects.filter(gstin_no=gstin_no).exists():
            errors['gstin_no'] = 'GSTIN already exists.'

        if mobile_no:
            if not mobile_no.isdigit():
                errors['mobile_no'] = 'Mobile number should contain digits only.'
            elif len(mobile_no) != 10:
                errors['mobile_no'] = 'Mobile number must be exactly 10 digits.'

        if phone_no:
            if not phone_no.isdigit():
                errors['phone_no'] = 'Phone number should contain digits only.'
            elif len(phone_no) != 10:
                errors['phone_no'] = 'Phone number must be exactly 10 digits.'

        if gstin_no:
            if len(gstin_no) != 15:
                errors['gstin_no'] = 'GSTIN number must be exactly 15 characters.'
            elif not gstin_no.isalnum():
                errors['gstin_no'] = 'GSTIN number must contain only alphanumeric characters.'

        if not errors:
            Company.objects.create(
                company_group=company_group,
                address=address,
                billing_name=billing_name,
                billing_address=billing_address,
                mobile_no=mobile_no,
                phone_no=phone_no,
                gstin_no=gstin_no,
                status=status,
                logo=logo,
            )
            messages.success(request, 'Company created successfully.')
            return redirect('master:company_list')

        return render(
            request,
            'master/company_creation/create.html',
            {
                'errors': errors,
                'submitted': submitted,
            },
        )

    return render(request, 'master/company_creation/create.html', {
        'errors': {},
        'submitted': {},
    })


@permission_required('master.change_company', raise_exception=True)
def company_edit(request, pk):
    company = get_object_or_404(Company, pk=pk)

    if request.method == 'POST':
        submitted = request.POST.dict()
        errors = {}

        company_group = submitted.get('company_group', '').strip()
        billing_name = submitted.get('billing_name', '').strip()
        mobile_no = submitted.get('mobile_no', '').strip()
        phone_no = submitted.get('phone_no', '').strip()
        address = submitted.get('address', '').strip()
        billing_address = submitted.get('billing_address', '').strip()
        gstin_no = submitted.get('gstin_no', '').strip()
        status = submitted.get('status', Company.STATUS_ACTIVE)
        logo = request.FILES.get('logo')

        required_fields = [
            ('company_group', company_group),
            ('billing_name', billing_name),
            ('address', address),
            ('billing_address', billing_address),
            ('mobile_no', mobile_no),
            ('gstin_no', gstin_no),
        ]
        for field, value in required_fields:
            if not value:
                errors[field] = 'This field is required.'

        if gstin_no and Company.objects.filter(gstin_no=gstin_no).exclude(pk=company.pk).exists():
            errors['gstin_no'] = 'GSTIN already exists.'

        if mobile_no:
            if not mobile_no.isdigit():
                errors['mobile_no'] = 'Mobile number should contain digits only.'
            elif len(mobile_no) != 10:
                errors['mobile_no'] = 'Mobile number must be exactly 10 digits.'

        if phone_no:
            if not phone_no.isdigit():
                errors['phone_no'] = 'Phone number should contain digits only.'
            elif len(phone_no) != 10:
                errors['phone_no'] = 'Phone number must be exactly 10 digits.'

        if gstin_no:
            if len(gstin_no) != 15:
                errors['gstin_no'] = 'GSTIN number must be exactly 15 characters.'
            elif not gstin_no.isalnum():
                errors['gstin_no'] = 'GSTIN number must contain only alphanumeric characters.'

        if not errors:
            company.company_group = company_group
            company.address = address
            company.billing_name = billing_name
            company.billing_address = billing_address
            company.mobile_no = mobile_no
            company.phone_no = phone_no
            company.gstin_no = gstin_no
            company.status = status
            if logo:
                company.logo = logo
            company.save()
            messages.success(request, 'Company updated successfully.')
            return redirect('master:company_list')

        return render(
            request,
            'master/company_creation/editForm.html',
            {
                'errors': errors,
                'submitted': submitted,
                'company': company,
            },
        )

    return render(request, 'master/company_creation/editForm.html', {
        'company': company,
        'errors': {},
        'submitted': {},
    })


@permission_required('master.delete_company', raise_exception=True)
def company_delete(request, pk):
    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        company.delete()
        messages.success(request, 'Company deleted successfully.')
        return redirect('master:company_list')

    return render(request, 'master/company_creation/confirm_delete.html', {'company': company})


@permission_required('master.view_company', raise_exception=True)
def company_list1(request):
    companies = Company.objects.order_by('-created_at')
    return render(request, 'master/company_creation/list1.html', {'companies': companies})


# ==================== Legacy Master Templates ====================

@permission_required('master.view_additiondeduction', raise_exception=True)
def addition_list(request):
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    additions = AdditionDeduction.objects.order_by('type', 'name')
    if search_query:
        additions = additions.filter(
            Q(type__icontains=search_query)
            | Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    paginator = Paginator(additions, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'additions': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'per_page': str(per_page_value),
        'per_page_value': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_additions': paginator.count,
    }
    return render(request, 'master/addition_creation/list.html', context)


@permission_required('master.add_additiondeduction', raise_exception=True)
def addition_create(request):
    values = {
        'type': '',
        'name': '',
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values['type'] = request.POST.get('type', '').strip()
        values['name'] = request.POST.get('name', '').strip()
        values['description'] = request.POST.get('description', '').strip()

        if not values['type']:
            errors['type'] = 'Type selection is required.'
        elif values['type'] not in dict(AdditionDeduction.TYPE_CHOICES):
            errors['type'] = 'Invalid type selected.'

        if not values['name']:
            errors['name'] = 'Name is required.'
        elif AdditionDeduction.objects.filter(
            type=values['type'],
            name__iexact=values['name']
        ).exists():
            errors['name'] = f'Name already exists for {values["type"]}.'

        if not errors:
            AdditionDeduction.objects.create(
                type=values['type'],
                name=values['name'],
                description=values['description'],
            )
            messages.success(request, 'Addition/Deduction created successfully.')
            return redirect('master:addition_list')

    context = {
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:addition_list'),
        'type_choices': AdditionDeduction.TYPE_CHOICES,
    }
    return render(request, 'master/addition_creation/create.html', context)


@permission_required('master.change_additiondeduction', raise_exception=True)
def addition_edit(request, pk):
    addition = get_object_or_404(AdditionDeduction, pk=pk)
    values = {
        'type': addition.type,
        'name': addition.name,
        'description': addition.description or '',
    }
    errors = {}

    if request.method == 'POST':
        values['type'] = request.POST.get('type', '').strip()
        values['name'] = request.POST.get('name', '').strip()
        values['description'] = request.POST.get('description', '').strip()

        if not values['type']:
            errors['type'] = 'Type selection is required.'
        elif values['type'] not in dict(AdditionDeduction.TYPE_CHOICES):
            errors['type'] = 'Invalid type selected.'

        if not values['name']:
            errors['name'] = 'Name is required.'
        elif AdditionDeduction.objects.filter(
            type=values['type'],
            name__iexact=values['name']
        ).exclude(pk=addition.pk).exists():
            errors['name'] = f'Name already exists for {values["type"]}.'

        if not errors:
            addition.type = values['type']
            addition.name = values['name']
            addition.description = values['description']
            addition.save()
            messages.success(request, 'Addition/Deduction updated successfully.')
            return redirect('master:addition_list')

    context = {
        'addition': addition,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:addition_list'),
        'type_choices': AdditionDeduction.TYPE_CHOICES,
    }
    return render(request, 'master/addition_creation/edit.html', context)


@permission_required('master.delete_additiondeduction', raise_exception=True)
def addition_delete(request, pk):
    addition = get_object_or_404(AdditionDeduction, pk=pk)
    if request.method == 'POST':
        addition.delete()
        messages.success(request, 'Addition/Deduction deleted successfully.')
        return redirect('master:addition_list')

    context = {
        'addition': addition,
    }
    return render(request, 'master/addition_creation/confirm_delete.html', context)


@permission_required('master.view_assettype', raise_exception=True)
def asset_list(request):
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    assets = AssetType.objects.order_by('name')
    if search_query:
        assets = assets.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )

    paginator = Paginator(assets, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'assets': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'per_page': str(per_page_value),
        'per_page_value': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_assets': paginator.count,
    }
    return render(request, 'master/asset_creation/list.html', context)


@permission_required('master.view_employeeassetassignment', raise_exception=True)
def asset_create_list(request):
    """List all asset assignments with filters."""
    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()
    site_id = request.GET.get('site_id', '').strip()
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('search', '').strip()
    page_number = request.GET.get('page')
    
    # Get all asset assignments with related objects
    asset_assignments = EmployeeAssetAssignment.objects.select_related(
        'employee', 'employee__company', 'asset_type'
    ).all()
    
    # Filter by date range (using created_at)
    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            asset_assignments = asset_assignments.filter(created_at__date__gte=from_date_obj)
        except ValueError:
            pass
    
    if to_date:
        try:
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            asset_assignments = asset_assignments.filter(created_at__date__lte=to_date_obj)
        except ValueError:
            pass
    
    # Filter by site (if employee has site relationship - for now skip, can be added later)
    # if site_id:
    #     try:
    #         asset_assignments = asset_assignments.filter(employee__site_id=int(site_id))
    #     except (ValueError, TypeError):
    #         pass
    
    # Search filter
    if search_query:
        asset_assignments = asset_assignments.filter(
            Q(asset_name__icontains=search_query) |
            Q(asset_type__name__icontains=search_query) |
            Q(serial_no__icontains=search_query) |
            Q(employee__staff_name__icontains=search_query) |
            Q(employee__staff_id__icontains=search_query)
        )
    
    # Pagination
    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10
    
    paginator = Paginator(asset_assignments.order_by('-created_at'), per_page_value)
    page_obj = paginator.get_page(page_number)
    
    # Get sites and other data for filters
    sites = Site.objects.all().order_by('name')
    
    # Calculate issued and returned quantities
    for assignment in page_obj:
        assignment.issued_qty = assignment.quantity if assignment.status == EmployeeAssetAssignment.STATUS_ISSUED else 0
        assignment.returned_qty = assignment.quantity if assignment.status == EmployeeAssetAssignment.STATUS_RETURNED else 0
        # Generate asset ID (AST-{id})
        assignment.asset_id_display = f'AST-{assignment.id:07d}'
    
    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''
    
    context = {
        'asset_assignments': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'sites': sites,
        'from_date': from_date,
        'to_date': to_date,
        'site_id': site_id,
        'search_query': search_query,
        'per_page': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
    }
    return render(request, 'master/assetCreate_creation/list.html', context)


@permission_required('master.add_employeeassetassignment', raise_exception=True)
def asset_create_create(request):
    """Create new asset assignment."""
    sites = Site.objects.all().order_by('name')
    employees = Employee.objects.all().order_by('staff_name')
    asset_types = AssetType.objects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        # Get all items from form data (items[0][field], items[1][field], etc.)
        items_data = {}
        for key, value in request.POST.items():
            if key.startswith('items[') and ']' in key:
                # Parse items[0][field] format
                try:
                    # Extract index and field name
                    key_parts = key.replace('items[', '').split('][')
                    if len(key_parts) == 2:
                        item_index = key_parts[0]
                        field_name = key_parts[1].replace(']', '')
                        
                        if item_index not in items_data:
                            items_data[item_index] = {}
                        items_data[item_index][field_name] = value
                except Exception:
                    continue
        
        # Validate and save each item
        saved_count = 0
        errors = []
        
        if not items_data:
            messages.error(request, 'No items found to save. Please add at least one item to the sub-list.')
        else:
            # Sort by index to maintain order
            sorted_indices = sorted(items_data.keys(), key=lambda x: int(x) if x.isdigit() else 0)
            
            for item_index in sorted_indices:
                item = items_data[item_index]
                
                # Get required fields
                staff_id = item.get('staff_id', '').strip()
                asset_type_id = item.get('asset_type_id', '').strip()
                serial_no = item.get('serial_no', '').strip()
                quantity = item.get('quantity', '1').strip()
                status = item.get('status', '').strip() or EmployeeAssetAssignment.STATUS_ISSUED
                description = item.get('description', '').strip()
                
                # Validate required fields
                if not staff_id:
                    errors.append(f'Item {int(item_index) + 1}: Staff name is required.')
                    continue
                
                if not asset_type_id:
                    errors.append(f'Item {int(item_index) + 1}: Asset type is required.')
                    continue
                
                if not serial_no:
                    errors.append(f'Item {int(item_index) + 1}: Serial No is required.')
                    continue
                
                # Get Employee object
                try:
                    employee = Employee.objects.get(pk=staff_id)
                except Employee.DoesNotExist:
                    errors.append(f'Item {int(item_index) + 1}: Invalid employee selected.')
                    continue
                
                # Get AssetType object (ForeignKey reference)
                try:
                    asset_type = AssetType.objects.get(pk=asset_type_id)
                except AssetType.DoesNotExist:
                    errors.append(f'Item {int(item_index) + 1}: Invalid asset type selected.')
                    continue
                
                # Parse quantity
                try:
                    quantity_int = max(int(quantity or '1'), 1)
                except (ValueError, TypeError):
                    quantity_int = 1
                
                # Validate status
                valid_statuses = [choice[0] for choice in EmployeeAssetAssignment.STATUS_CHOICES]
                if status not in valid_statuses:
                    status = EmployeeAssetAssignment.STATUS_ISSUED
                
                # Get description if provided
                description = item.get('description', '').strip()
                
                # Create EmployeeAssetAssignment record with AssetType ForeignKey
                try:
                    EmployeeAssetAssignment.objects.create(
                        employee=employee,
                        asset_type=asset_type,  # ForeignKey to AssetType
                        asset_name=asset_type.name,  # Keep for backward compatibility
                        serial_no=serial_no,
                        quantity=quantity_int,
                        status=status,
                        description=description,
                    )
                    saved_count += 1
                except Exception as e:
                    errors.append(f'Item {int(item_index) + 1}: Error saving - {str(e)}')
            
            # Show success or error messages
            if errors:
                for error in errors:
                    messages.error(request, error)
                if saved_count > 0:
                    messages.warning(request, f'{saved_count} item(s) saved successfully, but some items had errors.')
            else:
                if saved_count > 0:
                    messages.success(request, f'{saved_count} asset assignment(s) created successfully.')
                else:
                    messages.error(request, 'No items were saved.')
            
            # Redirect to list page
            return redirect('master:asset_create_list')
    
    # GET request - show form
    context = {
        'sites': sites,
        'employees': employees,
        'asset_types': asset_types,
    }
    return render(request, 'master/assetCreate_creation/create.html', context)


@permission_required('master.change_employeeassetassignment', raise_exception=True)
def asset_create_edit(request, pk):
    """Edit existing asset assignment."""
    assignment = get_object_or_404(EmployeeAssetAssignment, pk=pk)
    sites = Site.objects.all().order_by('name')
    employees = Employee.objects.all().order_by('staff_name')
    asset_types = AssetType.objects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        # Get form data
        staff_id = request.POST.get('staff_name', '').strip()
        asset_type_id = request.POST.get('asset_type', '').strip()
        serial_no = request.POST.get('serial_no', '').strip()
        quantity = request.POST.get('quantity', '1').strip()
        status = request.POST.get('status', '').strip() or EmployeeAssetAssignment.STATUS_ISSUED
        
        # Validate required fields
        errors = []
        
        if not staff_id:
            errors.append('Staff name is required.')
        
        if not asset_type_id:
            errors.append('Asset type is required.')
        
        if not serial_no:
            errors.append('Serial No is required.')
        
        # Get Employee object
        try:
            employee = Employee.objects.get(pk=staff_id)
        except Employee.DoesNotExist:
            errors.append('Invalid employee selected.')
            employee = assignment.employee  # Keep existing if invalid
        
        # Get AssetType object (ForeignKey reference)
        try:
            asset_type = AssetType.objects.get(pk=asset_type_id)
        except AssetType.DoesNotExist:
            errors.append('Invalid asset type selected.')
            asset_type = assignment.asset_type  # Keep existing if invalid
        
        # Parse quantity
        try:
            quantity_int = max(int(quantity or '1'), 1)
        except (ValueError, TypeError):
            quantity_int = assignment.quantity  # Keep existing if invalid
        
        # Validate status
        valid_statuses = [choice[0] for choice in EmployeeAssetAssignment.STATUS_CHOICES]
        if status not in valid_statuses:
            status = assignment.status  # Keep existing if invalid
        
        # If no errors, update and save
        if not errors:
            assignment.employee = employee
            assignment.asset_type = asset_type  # ForeignKey to AssetType
            assignment.asset_name = asset_type.name if asset_type else assignment.asset_name  # Keep for backward compatibility
            assignment.serial_no = serial_no
            assignment.quantity = quantity_int
            assignment.status = status
            # Get description if provided
            description = request.POST.get('description', '').strip()
            if description:
                assignment.description = description
            assignment.save()
            
            messages.success(request, 'Asset assignment updated successfully.')
            return redirect('master:asset_create_list')
        else:
            # Show errors
            for error in errors:
                messages.error(request, error)
    
    # GET request or errors - show form
    context = {
        'assignment': assignment,
        'sites': sites,
        'employees': employees,
        'asset_types': asset_types,
    }
    return render(request, 'master/assetCreate_creation/edit.html', context)


@permission_required('master.delete_employeeassetassignment', raise_exception=True)
@require_POST
def asset_create_delete(request, pk):
    """Delete asset assignment."""
    assignment = get_object_or_404(EmployeeAssetAssignment, pk=pk)
    employee_name = assignment.employee.staff_name
    asset_name = assignment.asset_name_display
    assignment.delete()
    messages.success(request, f'Asset assignment for {employee_name} - {asset_name} deleted successfully.')
    return redirect('master:asset_create_list')


@permission_required('master.view_employeeassetassignment', raise_exception=True)
def asset_create_print(request, pk):
    """Print asset assignment."""
    assignment = get_object_or_404(
        EmployeeAssetAssignment.objects.select_related('employee', 'employee__company', 'asset_type'),
        pk=pk
    )
    
    # Generate asset ID
    assignment.asset_id_display = f'AST-{assignment.id:07d}'
    assignment.issued_qty = assignment.quantity if assignment.status == EmployeeAssetAssignment.STATUS_ISSUED else 0
    assignment.returned_qty = assignment.quantity if assignment.status == EmployeeAssetAssignment.STATUS_RETURNED else 0
    
    context = {
        'assignment': assignment,
    }
    return render(request, 'master/assetCreate_creation/print.html', context)


@permission_required('master.add_employeeassetassignment', raise_exception=True)
def asset_create(request):
    values = {
        'name': '',
        'description': '',
        'is_active': True,
    }
    errors = {}

    if request.method == 'POST':
        values['name'] = request.POST.get('name', '').strip()
        values['description'] = request.POST.get('description', '').strip()
        values['is_active'] = request.POST.get('is_active') == 'on'

        if not values['name']:
            errors['name'] = 'Asset type is required.'
        elif AssetType.objects.filter(name__iexact=values['name']).exists():
            errors['name'] = 'Asset type already exists.'

        if not errors:
            AssetType.objects.create(
                name=values['name'],
                description=values['description'],
                is_active=values['is_active'],
            )
            messages.success(request, 'Asset type created successfully.')
            return redirect('master:asset_list')

    context = {
        'values': values,
        'errors': errors,
    }
    return render(request, 'master/asset_creation/create.html', context)


@permission_required('master.change_employeeassetassignment', raise_exception=True)
def asset_edit(request, pk):
    asset = get_object_or_404(AssetType, pk=pk)
    values = {
        'name': asset.name,
        'description': asset.description or '',
        'is_active': asset.is_active,
    }
    errors = {}

    if request.method == 'POST':
        values['name'] = request.POST.get('name', '').strip()
        values['description'] = request.POST.get('description', '').strip()
        values['is_active'] = request.POST.get('is_active') == 'on'

        if not values['name']:
            errors['name'] = 'Asset type is required.'
        elif (
            AssetType.objects.filter(name__iexact=values['name'])
            .exclude(pk=asset.pk)
            .exists()
        ):
            errors['name'] = 'Asset type already exists.'

        if not errors:
            asset.name = values['name']
            asset.description = values['description']
            asset.is_active = values['is_active']
            asset.save()
            messages.success(request, 'Asset type updated successfully.')
            return redirect('master:asset_list')

    context = {
        'asset': asset,
        'values': values,
        'errors': errors,
    }
    return render(request, 'master/asset_creation/edit.html', context)


@permission_required('master.delete_employeeassetassignment', raise_exception=True)
def asset_delete(request, pk):
    asset = get_object_or_404(AssetType, pk=pk)
    if request.method == 'POST':
        asset.delete()
        messages.success(request, 'Asset type deleted successfully.')
        return redirect('master:asset_list')

    context = {
        'asset': asset,
    }
    return render(request, 'master/asset_creation/confirm_delete.html', context)


@permission_required('master.view_department', raise_exception=True)
def department_list(request):
    departments = Department.objects.order_by('name')
    status_filter = request.GET.get('status', '').strip()
    search_query = request.GET.get('q', '').strip()
    per_page = request.GET.get('per_page', '').strip() or '10'
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    if status_filter:
        departments = departments.filter(status=status_filter)
    if search_query:
        departments = departments.filter(
            Q(name__icontains=search_query)
            | Q(status__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    paginator = Paginator(departments, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'departments': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'status_filter': status_filter,
        'search_query': search_query,
        'per_page': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
    }
    return render(request, 'master/department_creation/list.html', context)


@permission_required('master.add_department', raise_exception=True)
def department_create(request):
    values = {
        'name': '',
        'status': Department.STATUS_ACTIVE,
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values['name'] = request.POST.get('name', '').strip()
        values['status'] = request.POST.get('status', Department.STATUS_ACTIVE).strip() or Department.STATUS_ACTIVE
        values['description'] = request.POST.get('description', '').strip()

        if not values['name']:
            errors['name'] = 'Department name is required.'
        elif Department.objects.filter(name__iexact=values['name']).exists():
            errors['name'] = 'Department with this name already exists.'

        if values['status'] not in dict(Department.STATUS_CHOICES):
            errors['status'] = 'Invalid status selected.'

        if not errors:
            Department.objects.create(
                name=values['name'],
                status=values['status'],
                description=values['description'],
            )
            messages.success(request, 'Department created successfully.')
            return redirect('master:department_list')

    context = {
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:department_list'),
    }
    return render(request, 'master/department_creation/create.html', context)


@permission_required('master.change_department', raise_exception=True)
def department_edit(request, pk):
    department = get_object_or_404(Department, pk=pk)
    values = {
        'name': department.name,
        'status': department.status,
        'description': department.description or '',
    }
    errors = {}

    if request.method == 'POST':
        values['name'] = request.POST.get('name', '').strip()
        values['status'] = request.POST.get('status', Department.STATUS_ACTIVE).strip() or Department.STATUS_ACTIVE
        values['description'] = request.POST.get('description', '').strip()

        if not values['name']:
            errors['name'] = 'Department name is required.'

        if (
            Department.objects.filter(name__iexact=values['name'])
            .exclude(pk=department.pk)
            .exists()
        ):
            errors['name'] = 'Department with this name already exists.'

        if values['status'] not in dict(Department.STATUS_CHOICES):
            errors['status'] = 'Invalid status selected.'

        if not errors:
            department.name = values['name']
            department.status = values['status']
            department.description = values['description']
            department.save()
            messages.success(request, 'Department updated successfully.')
            return redirect('master:department_list')

    context = {
        'department': department,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:department_list'),
    }
    return render(request, 'master/department_creation/edit.html', context)


@permission_required('master.delete_department', raise_exception=True)
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        department.delete()
        messages.success(request, 'Department deleted successfully.')
        return redirect('master:department_list')

    context = {
        'department': department,
    }
    return render(request, 'master/department_creation/confirm_delete.html', context)


@permission_required('master.view_designation', raise_exception=True)
def designation_list(request):
    designations = (
        Designation.objects.select_related('department')
        .order_by('department__name', 'name')
    )

    status_filter = request.GET.get('status', '').strip()
    department_filter = request.GET.get('department', '').strip()
    search_query = request.GET.get('q', '').strip()
    per_page = request.GET.get('per_page', '').strip() or '10'
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    if status_filter:
        designations = designations.filter(status=status_filter)

    if department_filter:
        designations = designations.filter(department_id=department_filter)

    if search_query:
        designations = designations.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(department__name__icontains=search_query)
        )

    paginator = Paginator(designations, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'designations': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'status_filter': status_filter,
        'department_filter': department_filter,
        'search_query': search_query,
        'per_page': per_page_value,
        'departments': Department.objects.order_by('name'),
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
    }
    return render(request, 'master/designation_creation/list.html', context)


@permission_required('master.add_designation', raise_exception=True)
def designation_create(request):
    departments = Department.objects.order_by('name')
    values = {
        'department': '',
        'name': '',
        'status': Designation.STATUS_ACTIVE,
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values['department'] = request.POST.get('department', '').strip()
        values['name'] = request.POST.get('name', '').strip()
        values['status'] = request.POST.get('status', Designation.STATUS_ACTIVE).strip() or Designation.STATUS_ACTIVE
        values['description'] = request.POST.get('description', '').strip()

        department_instance = None
        if not values['department']:
            errors['department'] = 'Department selection is required.'
        else:
            try:
                department_instance = departments.get(pk=int(values['department']))
            except (Department.DoesNotExist, ValueError, TypeError):
                errors['department'] = 'Invalid department selected.'

        if not values['name']:
            errors['name'] = 'Designation name is required.'
        elif department_instance and Designation.objects.filter(
            department=department_instance,
            name__iexact=values['name'],
        ).exists():
            errors['name'] = 'Designation with this name already exists in the selected department.'

        if values['status'] not in dict(Designation.STATUS_CHOICES):
            errors['status'] = 'Invalid status selected.'

        if not errors and department_instance:
            Designation.objects.create(
                department=department_instance,
                name=values['name'],
                status=values['status'],
                description=values['description'],
            )
            messages.success(request, 'Designation created successfully.')
            return redirect('master:designation_list')

    context = {
        'departments': departments,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:designation_list'),
    }
    return render(request, 'master/designation_creation/create.html', context)


@permission_required('master.change_designation', raise_exception=True)
def designation_edit(request, pk):
    designation = get_object_or_404(Designation.objects.select_related('department'), pk=pk)
    departments = Department.objects.order_by('name')
    values = {
        'department': str(designation.department_id),
        'name': designation.name,
        'status': designation.status,
        'description': designation.description or '',
    }
    errors = {}

    if request.method == 'POST':
        values['department'] = request.POST.get('department', '').strip()
        values['name'] = request.POST.get('name', '').strip()
        values['status'] = request.POST.get('status', Designation.STATUS_ACTIVE).strip() or Designation.STATUS_ACTIVE
        values['description'] = request.POST.get('description', '').strip()

        department_instance = None
        if not values['department']:
            errors['department'] = 'Department selection is required.'
        else:
            try:
                department_instance = departments.get(pk=int(values['department']))
            except (Department.DoesNotExist, ValueError, TypeError):
                errors['department'] = 'Invalid department selected.'

        if not values['name']:
            errors['name'] = 'Designation name is required.'
        elif department_instance and Designation.objects.filter(
            department=department_instance,
            name__iexact=values['name'],
        ).exclude(pk=designation.pk).exists():
            errors['name'] = 'Designation with this name already exists in the selected department.'

        if values['status'] not in dict(Designation.STATUS_CHOICES):
            errors['status'] = 'Invalid status selected.'

        if not errors and department_instance:
            designation.department = department_instance
            designation.name = values['name']
            designation.status = values['status']
            designation.description = values['description']
            designation.save()
            messages.success(request, 'Designation updated successfully.')
            return redirect('master:designation_list')

    context = {
        'designation': designation,
        'departments': departments,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:designation_list'),
    }
    return render(request, 'master/designation_creation/edit.html', context)


@permission_required('master.delete_designation', raise_exception=True)
def designation_delete(request, pk):
    designation = get_object_or_404(Designation.objects.select_related('department'), pk=pk)
    if request.method == 'POST':
        designation.delete()
        messages.success(request, 'Designation deleted successfully.')
        return redirect('master:designation_list')

    context = {
        'designation': designation,
    }
    return render(request, 'master/designation_creation/confirm_delete.html', context)


@permission_required('master.view_degree', raise_exception=True)
def degree_list(request):
    degrees = Degree.objects.order_by('education_type', 'name')

    search_query = request.GET.get('q', '').strip()
    per_page = request.GET.get('per_page', '').strip() or '10'
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    if search_query:
        degrees = degrees.filter(
            Q(name__icontains=search_query)
            | Q(education_type__icontains=search_query)
        )

    paginator = Paginator(degrees, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'degrees': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'search_query': search_query,
        'per_page': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
    }
    return render(request, 'master/degree_creation/list.html', context)


@permission_required('master.add_degree', raise_exception=True)
def degree_create(request):
    values = {
        'education_type': Degree.TYPE_UG,
        'name': '',
    }
    errors = {}

    if request.method == 'POST':
        values['education_type'] = request.POST.get('education_type', Degree.TYPE_UG).strip() or Degree.TYPE_UG
        values['name'] = request.POST.get('name', '').strip()

        if not values['name']:
            errors['name'] = 'Degree name is required.'
        elif Degree.objects.filter(name__iexact=values['name']).exists():
            errors['name'] = 'Degree with this name already exists.'

        if values['education_type'] not in dict(Degree.EDUCATION_TYPE_CHOICES):
            errors['education_type'] = 'Invalid education type selected.'

        if not errors:
            Degree.objects.create(
                education_type=values['education_type'],
                name=values['name'],
            )
            messages.success(request, 'Degree created successfully.')
            return redirect('master:degree_list')

    context = {
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:degree_list'),
        'education_choices': Degree.EDUCATION_TYPE_CHOICES,
    }
    return render(request, 'master/degree_creation/create.html', context)


@permission_required('master.change_degree', raise_exception=True)
def degree_edit(request, pk):
    degree = get_object_or_404(Degree, pk=pk)
    values = {
        'education_type': degree.education_type,
        'name': degree.name,
    }
    errors = {}

    if request.method == 'POST':
        values['education_type'] = request.POST.get('education_type', Degree.TYPE_UG).strip() or Degree.TYPE_UG
        values['name'] = request.POST.get('name', '').strip()

        if not values['name']:
            errors['name'] = 'Degree name is required.'
        elif Degree.objects.filter(name__iexact=values['name']).exclude(pk=degree.pk).exists():
            errors['name'] = 'Degree with this name already exists.'

        if values['education_type'] not in dict(Degree.EDUCATION_TYPE_CHOICES):
            errors['education_type'] = 'Invalid education type selected.'

        if not errors:
            degree.education_type = values['education_type']
            degree.name = values['name']
            degree.save()
            messages.success(request, 'Degree updated successfully.')
            return redirect('master:degree_list')

    context = {
        'degree': degree,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:degree_list'),
        'education_choices': Degree.EDUCATION_TYPE_CHOICES,
    }
    return render(request, 'master/degree_creation/edit.html', context)


@permission_required('master.delete_degree', raise_exception=True)
def degree_delete(request, pk):
    degree = get_object_or_404(Degree, pk=pk)
    if request.method == 'POST':
        degree.delete()
        messages.success(request, 'Degree deleted successfully.')
        return redirect('master:degree_list')

    context = {
        'degree': degree,
    }
    return render(request, 'master/degree_creation/confirm_delete.html', context)


@permission_required('master.view_expensetype', raise_exception=True)
def expense_list(request):
    return render(request, 'master/expense_creation/list.html')


@permission_required('master.view_subexpense', raise_exception=True)
def sub_expense_list(request):
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    sub_expenses = SubExpense.objects.select_related('expense_type').order_by('-entry_date', 'expense_type__name', 'name')
    if search_query:
        sub_expenses = sub_expenses.filter(
            Q(name__icontains=search_query)
            | Q(expense_type__name__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    paginator = Paginator(sub_expenses, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'sub_expenses': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'per_page': str(per_page_value),
        'per_page_value': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_sub_expenses': paginator.count,
    }
    return render(request, 'master/subExpense_creation/list.html', context)


@permission_required('master.add_subexpense', raise_exception=True)
def sub_expense_create(request):
    expense_types = ExpenseType.objects.order_by('name')
    values = {
        'entry_date': '',
        'expense_type': '',
        'name': '',
        'image_required': False,
        'image_type': '',
        'meter_type': '',
        'status': SubExpense.STATUS_ACTIVE,
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values['entry_date'] = request.POST.get('entry_date', '').strip()
        values['expense_type'] = request.POST.get('expense_type', '').strip()
        values['name'] = request.POST.get('name', '').strip()
        values['image_required'] = request.POST.get('image_required') == 'Yes'
        values['image_type'] = request.POST.get('image_type', '').strip()
        values['meter_type'] = request.POST.get('meter_type', '').strip()
        values['status'] = request.POST.get('status', SubExpense.STATUS_ACTIVE).strip() or SubExpense.STATUS_ACTIVE
        values['description'] = request.POST.get('description', '').strip()

        if not values['entry_date']:
            errors['entry_date'] = 'Entry date is required.'
        else:
            try:
                datetime.strptime(values['entry_date'], '%Y-%m-%d').date()
            except ValueError:
                errors['entry_date'] = 'Invalid date format.'

        expense_type_instance = None
        if not values['expense_type']:
            errors['expense_type'] = 'Expense type is required.'
        else:
            try:
                expense_type_instance = expense_types.get(pk=int(values['expense_type']))
            except (ExpenseType.DoesNotExist, ValueError, TypeError):
                errors['expense_type'] = 'Invalid expense type selected.'

        if not values['name']:
            errors['name'] = 'Sub expense type is required.'
        elif expense_type_instance and SubExpense.objects.filter(
            expense_type=expense_type_instance,
            name__iexact=values['name']
        ).exists():
            errors['name'] = f'Sub expense type already exists for {expense_type_instance.name}.'

        if values['image_required'] and not values['image_type']:
            errors['image_type'] = 'Image type is required when image is required.'

        if values['status'] not in dict(SubExpense.STATUS_CHOICES):
            errors['status'] = 'Invalid status selected.'

        if not errors and expense_type_instance:
            SubExpense.objects.create(
                entry_date=values['entry_date'],
                expense_type=expense_type_instance,
                name=values['name'],
                image_required=values['image_required'],
                image_type=values['image_type'] if values['image_required'] else '',
                meter_type=values['meter_type'],
                status=values['status'],
                description=values['description'],
            )
            messages.success(request, 'Sub expense created successfully.')
            return redirect('master:sub_expense_list')

    context = {
        'expense_types': expense_types,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:sub_expense_list'),
        'status_choices': SubExpense.STATUS_CHOICES,
        'image_type_choices': SubExpense.IMAGE_TYPE_CHOICES,
        'meter_type_choices': SubExpense.METER_TYPE_CHOICES,
    }
    return render(request, 'master/subExpense_creation/create.html', context)


@permission_required('master.change_subexpense', raise_exception=True)
def sub_expense_edit(request, pk):
    sub_expense = get_object_or_404(SubExpense.objects.select_related('expense_type'), pk=pk)
    expense_types = ExpenseType.objects.order_by('name')
    values = {
        'entry_date': sub_expense.entry_date.isoformat(),
        'expense_type': str(sub_expense.expense_type_id),
        'name': sub_expense.name,
        'image_required': sub_expense.image_required,
        'image_type': sub_expense.image_type or '',
        'meter_type': sub_expense.meter_type or '',
        'status': sub_expense.status,
        'description': sub_expense.description or '',
    }
    errors = {}

    if request.method == 'POST':
        values['entry_date'] = request.POST.get('entry_date', '').strip()
        values['expense_type'] = request.POST.get('expense_type', '').strip()
        values['name'] = request.POST.get('name', '').strip()
        values['image_required'] = request.POST.get('image_required') == 'Yes'
        values['image_type'] = request.POST.get('image_type', '').strip()
        values['meter_type'] = request.POST.get('meter_type', '').strip()
        values['status'] = request.POST.get('status', SubExpense.STATUS_ACTIVE).strip() or SubExpense.STATUS_ACTIVE
        values['description'] = request.POST.get('description', '').strip()

        if not values['entry_date']:
            errors['entry_date'] = 'Entry date is required.'
        else:
            try:
                datetime.strptime(values['entry_date'], '%Y-%m-%d').date()
            except ValueError:
                errors['entry_date'] = 'Invalid date format.'

        expense_type_instance = None
        if not values['expense_type']:
            errors['expense_type'] = 'Expense type is required.'
        else:
            try:
                expense_type_instance = expense_types.get(pk=int(values['expense_type']))
            except (ExpenseType.DoesNotExist, ValueError, TypeError):
                errors['expense_type'] = 'Invalid expense type selected.'

        if not values['name']:
            errors['name'] = 'Sub expense type is required.'
        elif expense_type_instance and SubExpense.objects.filter(
            expense_type=expense_type_instance,
            name__iexact=values['name']
        ).exclude(pk=sub_expense.pk).exists():
            errors['name'] = f'Sub expense type already exists for {expense_type_instance.name}.'

        if values['image_required'] and not values['image_type']:
            errors['image_type'] = 'Image type is required when image is required.'

        if values['status'] not in dict(SubExpense.STATUS_CHOICES):
            errors['status'] = 'Invalid status selected.'

        if not errors and expense_type_instance:
            sub_expense.entry_date = values['entry_date']
            sub_expense.expense_type = expense_type_instance
            sub_expense.name = values['name']
            sub_expense.image_required = values['image_required']
            sub_expense.image_type = values['image_type'] if values['image_required'] else ''
            sub_expense.meter_type = values['meter_type']
            sub_expense.status = values['status']
            sub_expense.description = values['description']
            sub_expense.save()
            messages.success(request, 'Sub expense updated successfully.')
            return redirect('master:sub_expense_list')

    context = {
        'sub_expense': sub_expense,
        'expense_types': expense_types,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:sub_expense_list'),
        'status_choices': SubExpense.STATUS_CHOICES,
        'image_type_choices': SubExpense.IMAGE_TYPE_CHOICES,
        'meter_type_choices': SubExpense.METER_TYPE_CHOICES,
    }
    return render(request, 'master/subExpense_creation/edit.html', context)


@permission_required('master.delete_subexpense', raise_exception=True)
def sub_expense_delete(request, pk):
    sub_expense = get_object_or_404(SubExpense.objects.select_related('expense_type'), pk=pk)
    if request.method == 'POST':
        sub_expense.delete()
        messages.success(request, 'Sub expense deleted successfully.')
        return redirect('master:sub_expense_list')

    context = {
        'sub_expense': sub_expense,
    }
    return render(request, 'master/subExpense_creation/confirm_delete.html', context)


@permission_required('master.view_site', raise_exception=True)
def site_list(request):
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    sites = Site.objects.order_by('name')
    if search_query:
        sites = sites.filter(
            Q(name__icontains=search_query)
            | Q(address__icontains=search_query)
            | Q(city__icontains=search_query)
            | Q(state__icontains=search_query)
        )

    paginator = Paginator(sites, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'sites': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'per_page': str(per_page_value),
        'per_page_value': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_sites': paginator.count,
    }
    return render(request, 'master/site_creation/list.html', context)


@permission_required('master.add_site', raise_exception=True)
def site_create(request):
    values = {
        'name': '',
        'address': '',
        'city': '',
        'state': '',
        'latitude': '',
        'longitude': '',
    }
    errors = {}

    if request.method == 'POST':
        values['name'] = request.POST.get('name', '').strip()
        values['address'] = request.POST.get('address', '').strip()
        values['city'] = request.POST.get('city', '').strip()
        values['state'] = request.POST.get('state', '').strip()
        values['latitude'] = request.POST.get('latitude', '').strip()
        values['longitude'] = request.POST.get('longitude', '').strip()

        if not values['name']:
            errors['name'] = 'Site name is required.'
        elif Site.objects.filter(name__iexact=values['name']).exists():
            errors['name'] = 'Site with this name already exists.'

        if values['latitude']:
            try:
                float(values['latitude'])
            except ValueError:
                errors['latitude'] = 'Invalid latitude value.'

        if values['longitude']:
            try:
                float(values['longitude'])
            except ValueError:
                errors['longitude'] = 'Invalid longitude value.'

        if not errors:
            Site.objects.create(
                name=values['name'],
                address=values['address'],
                city=values['city'],
                state=values['state'],
                latitude=float(values['latitude']) if values['latitude'] else None,
                longitude=float(values['longitude']) if values['longitude'] else None,
            )
            messages.success(request, 'Site created successfully.')
            return redirect('master:site_list')

    context = {
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:site_list'),
    }
    return render(request, 'master/site_creation/create.html', context)


@permission_required('master.change_site', raise_exception=True)
def site_edit(request, pk):
    site = get_object_or_404(Site, pk=pk)
    values = {
        'name': site.name,
        'address': site.address or '',
        'city': site.city or '',
        'state': site.state or '',
        'latitude': str(site.latitude) if site.latitude else '',
        'longitude': str(site.longitude) if site.longitude else '',
    }
    errors = {}

    if request.method == 'POST':
        values['name'] = request.POST.get('name', '').strip()
        values['address'] = request.POST.get('address', '').strip()
        values['city'] = request.POST.get('city', '').strip()
        values['state'] = request.POST.get('state', '').strip()
        values['latitude'] = request.POST.get('latitude', '').strip()
        values['longitude'] = request.POST.get('longitude', '').strip()

        if not values['name']:
            errors['name'] = 'Site name is required.'
        elif Site.objects.filter(name__iexact=values['name']).exclude(pk=site.pk).exists():
            errors['name'] = 'Site with this name already exists.'

        if values['latitude']:
            try:
                float(values['latitude'])
            except ValueError:
                errors['latitude'] = 'Invalid latitude value.'

        if values['longitude']:
            try:
                float(values['longitude'])
            except ValueError:
                errors['longitude'] = 'Invalid longitude value.'

        if not errors:
            site.name = values['name']
            site.address = values['address']
            site.city = values['city']
            site.state = values['state']
            site.latitude = float(values['latitude']) if values['latitude'] else None
            site.longitude = float(values['longitude']) if values['longitude'] else None
            site.save()
            messages.success(request, 'Site updated successfully.')
            return redirect('master:site_list')

    context = {
        'site': site,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:site_list'),
    }
    return render(request, 'master/site_creation/edit.html', context)


@permission_required('master.delete_site', raise_exception=True)
def site_delete(request, pk):
    site = get_object_or_404(Site, pk=pk)
    if request.method == 'POST':
        site.delete()
        messages.success(request, 'Site deleted successfully.')
        return redirect('master:site_list')

    context = {
        'site': site,
    }
    return render(request, 'master/site_creation/confirm_delete.html', context)


@permission_required('master.view_plant', raise_exception=True)
def plant_list(request):
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    plants = Plant.objects.select_related('site').order_by('site__name', 'name')
    if search_query:
        plants = plants.filter(
            Q(name__icontains=search_query)
            | Q(site__name__icontains=search_query)
        )

    paginator = Paginator(plants, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'plants': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'per_page': str(per_page_value),
        'per_page_value': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_plants': paginator.count,
    }
    return render(request, 'master/plant_creation/list.html', context)


@permission_required('master.add_plant', raise_exception=True)
def plant_create(request):
    sites = Site.objects.order_by('name')
    values = {
        'site': '',
        'name': '',
    }
    errors = {}

    if request.method == 'POST':
        values['site'] = request.POST.get('site', '').strip()
        values['name'] = request.POST.get('name', '').strip()

        site_instance = None
        if not values['site']:
            errors['site'] = 'Site selection is required.'
        else:
            try:
                site_instance = sites.get(pk=int(values['site']))
            except (Site.DoesNotExist, ValueError, TypeError):
                errors['site'] = 'Invalid site selected.'

        if not values['name']:
            errors['name'] = 'Plant name is required.'
        elif site_instance and Plant.objects.filter(
            site=site_instance,
            name__iexact=values['name']
        ).exists():
            errors['name'] = f'Plant with this name already exists for {site_instance.name}.'

        if not errors and site_instance:
            Plant.objects.create(
                site=site_instance,
                name=values['name'],
            )
            messages.success(request, 'Plant created successfully.')
            return redirect('master:plant_list')

    context = {
        'sites': sites,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:plant_list'),
    }
    return render(request, 'master/plant_creation/create.html', context)


@permission_required('master.change_plant', raise_exception=True)
def plant_edit(request, pk):
    plant = get_object_or_404(Plant.objects.select_related('site'), pk=pk)
    sites = Site.objects.order_by('name')
    values = {
        'site': str(plant.site_id),
        'name': plant.name,
    }
    errors = {}

    if request.method == 'POST':
        values['site'] = request.POST.get('site', '').strip()
        values['name'] = request.POST.get('name', '').strip()

        site_instance = None
        if not values['site']:
            errors['site'] = 'Site selection is required.'
        else:
            try:
                site_instance = sites.get(pk=int(values['site']))
            except (Site.DoesNotExist, ValueError, TypeError):
                errors['site'] = 'Invalid site selected.'

        if not values['name']:
            errors['name'] = 'Plant name is required.'
        elif site_instance and Plant.objects.filter(
            site=site_instance,
            name__iexact=values['name']
        ).exclude(pk=plant.pk).exists():
            errors['name'] = f'Plant with this name already exists for {site_instance.name}.'

        if not errors and site_instance:
            plant.site = site_instance
            plant.name = values['name']
            plant.save()
            messages.success(request, 'Plant updated successfully.')
            return redirect('master:plant_list')

    context = {
        'plant': plant,
        'sites': sites,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:plant_list'),
    }
    return render(request, 'master/plant_creation/edit.html', context)


@permission_required('master.delete_plant', raise_exception=True)
def plant_delete(request, pk):
    plant = get_object_or_404(Plant.objects.select_related('site'), pk=pk)
    if request.method == 'POST':
        plant.delete()
        messages.success(request, 'Plant deleted successfully.')
        return redirect('master:plant_list')

    context = {
        'plant': plant,
    }
    return render(request, 'master/plant_creation/confirm_delete.html', context)


@permission_required('master.view_holiday', raise_exception=True)
def holidays_list(request):
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    holidays = Holiday.objects.order_by('-date', 'site_name')
    if search_query:
        holidays = holidays.filter(
            Q(site_name__icontains=search_query)
            | Q(holiday_type__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    paginator = Paginator(holidays, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'holidays': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'per_page': str(per_page_value),
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_holidays': paginator.count,
    }
    return render(request, 'master/holidays_creation/list.html', context)


@permission_required('master.add_holiday', raise_exception=True)
def holidays_create(request):
    values = {
        'date': '',
        'site_name': [],
        'holiday_type': '',
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values['date'] = request.POST.get('date', '').strip()
        site_names = [name.strip() for name in request.POST.getlist('site_name') if name.strip()]
        values['site_name'] = site_names
        values['holiday_type'] = request.POST.get('holiday_type', '').strip()
        values['description'] = request.POST.get('description', '').strip()

        if not values['date']:
            errors['date'] = 'Holiday date is required.'
        if not site_names:
            errors['site_name'] = 'Select at least one site.'
        if not values['holiday_type']:
            errors['holiday_type'] = 'Holiday type is required.'

        if not errors:
            Holiday.objects.create(
                date=values['date'],
                site_name=', '.join(site_names),
                holiday_type=values['holiday_type'],
                description=values['description'],
            )
            messages.success(request, 'Holiday created successfully.')
            return redirect('master:holidays_list')

    context = {
        'values': values,
        'errors': errors,
        'site_options': HOLIDAY_SITE_OPTIONS,
        'holiday_type_options': HOLIDAY_TYPE_OPTIONS,
    }
    return render(request, 'master/holidays_creation/create.html', context)


@permission_required('master.change_holiday', raise_exception=True)
def holidays_edit(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    values = {
        'date': holiday.date.isoformat(),
        'site_name': [name.strip() for name in holiday.site_name.split(',') if name.strip()],
        'holiday_type': holiday.holiday_type,
        'description': holiday.description or '',
    }
    errors = {}

    if request.method == 'POST':
        values['date'] = request.POST.get('date', '').strip()
        site_names = [name.strip() for name in request.POST.getlist('site_name') if name.strip()]
        values['site_name'] = site_names
        values['holiday_type'] = request.POST.get('holiday_type', '').strip()
        values['description'] = request.POST.get('description', '').strip()

        if not values['date']:
            errors['date'] = 'Holiday date is required.'
        if not site_names:
            errors['site_name'] = 'Select at least one site.'
        if not values['holiday_type']:
            errors['holiday_type'] = 'Holiday type is required.'

        if not errors:
            holiday.date = values['date']
            holiday.site_name = ', '.join(site_names)
            holiday.holiday_type = values['holiday_type']
            holiday.description = values['description']
            holiday.save()
            messages.success(request, 'Holiday updated successfully.')
            return redirect('master:holidays_list')

    context = {
        'holiday': holiday,
        'values': values,
        'errors': errors,
        'site_options': HOLIDAY_SITE_OPTIONS,
        'holiday_type_options': HOLIDAY_TYPE_OPTIONS,
    }
    return render(request, 'master/holidays_creation/edit.html', context)


@permission_required('master.delete_holiday', raise_exception=True)
def holidays_delete(request, pk):
    holiday = get_object_or_404(Holiday, pk=pk)
    if request.method == 'POST':
        holiday.delete()
        messages.success(request, 'Holiday deleted successfully.')
    return redirect('master:holidays_list')


@permission_required('master.view_leavetype', raise_exception=True)
def leave_list(request):
    per_page = request.GET.get('per_page', '').strip() or '10'
    search_query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    leave_types = LeaveType.objects.order_by('leave_type')
    if search_query:
        leave_types = leave_types.filter(
            Q(leave_type__icontains=search_query)
            | Q(short_name__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    paginator = Paginator(leave_types, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'leave_types': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'per_page': str(per_page_value),
        'per_page_value': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_leave_types': paginator.count,
    }
    return render(request, 'master/leave_creation/list.html', context)


@permission_required('master.add_leavetype', raise_exception=True)
def leave_create(request):
    values = {
        'leave_type': '',
        'short_name': '',
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values['leave_type'] = request.POST.get('leave_type', '').strip()
        values['short_name'] = request.POST.get('short_name', '').strip()
        values['description'] = request.POST.get('description', '').strip()

        if not values['leave_type']:
            errors['leave_type'] = 'Leave type is required.'
        elif LeaveType.objects.filter(leave_type__iexact=values['leave_type']).exists():
            errors['leave_type'] = 'Leave type already exists.'

        if not values['short_name']:
            errors['short_name'] = 'Short name is required.'
        elif LeaveType.objects.filter(short_name__iexact=values['short_name']).exists():
            errors['short_name'] = 'Short name already exists.'

        if not errors:
            LeaveType.objects.create(
                leave_type=values['leave_type'],
                short_name=values['short_name'],
                description=values['description'],
            )
            messages.success(request, 'Leave type created successfully.')
            return redirect('master:leave_list')

    context = {
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:leave_list'),
    }
    return render(request, 'master/leave_creation/create.html', context)


@permission_required('master.change_leavetype', raise_exception=True)
def leave_edit(request, pk):
    leave_type = get_object_or_404(LeaveType, pk=pk)
    values = {
        'leave_type': leave_type.leave_type,
        'short_name': leave_type.short_name,
        'description': leave_type.description or '',
    }
    errors = {}

    if request.method == 'POST':
        values['leave_type'] = request.POST.get('leave_type', '').strip()
        values['short_name'] = request.POST.get('short_name', '').strip()
        values['description'] = request.POST.get('description', '').strip()

        if not values['leave_type']:
            errors['leave_type'] = 'Leave type is required.'
        elif (
            LeaveType.objects.filter(leave_type__iexact=values['leave_type'])
            .exclude(pk=leave_type.pk)
            .exists()
        ):
            errors['leave_type'] = 'Leave type already exists.'

        if not values['short_name']:
            errors['short_name'] = 'Short name is required.'
        elif (
            LeaveType.objects.filter(short_name__iexact=values['short_name'])
            .exclude(pk=leave_type.pk)
            .exists()
        ):
            errors['short_name'] = 'Short name already exists.'

        if not errors:
            leave_type.leave_type = values['leave_type']
            leave_type.short_name = values['short_name']
            leave_type.description = values['description']
            leave_type.save()
            messages.success(request, 'Leave type updated successfully.')
            return redirect('master:leave_list')

    context = {
        'leave_type': leave_type,
        'values': values,
        'errors': errors,
        'cancel_url': reverse('master:leave_list'),
    }
    return render(request, 'master/leave_creation/edit.html', context)


@permission_required('master.delete_leavetype', raise_exception=True)
def leave_delete(request, pk):
    leave_type = get_object_or_404(LeaveType, pk=pk)
    if request.method == 'POST':
        leave_type.delete()
        messages.success(request, 'Leave type deleted successfully.')
        return redirect('master:leave_list')

    context = {
        'leave_type': leave_type,
    }
    return render(request, 'master/leave_creation/confirm_delete.html', context)


@permission_required('master.view_salarytype', raise_exception=True)
def salary_list(request):
    query = request.GET.get('q', '').strip()
    per_page = request.GET.get('per_page', '').strip() or '10'
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    salary_types = SalaryType.objects.order_by('name')
    if query:
        salary_types = salary_types.filter(name__icontains=query)

    paginator = Paginator(salary_types, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'salary_types': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'per_page': str(per_page_value),
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
        'total_salaries': paginator.count,
    }
    return render(request, 'master/salary_creation/list.html', context)


@permission_required('master.add_salarytype', raise_exception=True)
def salary_create(request):
    values = {
        'name': '',
        'description': '',
        'is_active': True,
    }
    errors = {}

    if request.method == 'POST':
        values.update({
            'name': request.POST.get('name', '').strip(),
            'description': request.POST.get('description', '').strip(),
            'is_active': request.POST.get('is_active') == 'on',
        })

        if not values['name']:
            errors['name'] = 'Salary type name is required.'
        elif SalaryType.objects.filter(name__iexact=values['name']).exists():
            errors['name'] = 'Salary type already exists.'

        if not errors:
            SalaryType.objects.create(**values)
            messages.success(request, 'Salary type created successfully.')
            return redirect('master:salary_list')

    context = {
        'values': values,
        'errors': errors,
    }
    return render(request, 'master/salary_creation/create.html', context)


@permission_required('master.change_salarytype', raise_exception=True)
def salary_edit(request, pk):
    salary_type = get_object_or_404(SalaryType, pk=pk)

    values = {
        'name': salary_type.name,
        'description': salary_type.description,
        'is_active': salary_type.is_active,
    }
    errors = {}

    if request.method == 'POST':
        values.update({
            'name': request.POST.get('name', '').strip(),
            'description': request.POST.get('description', '').strip(),
            'is_active': request.POST.get('is_active') == 'on',
        })

        if not values['name']:
            errors['name'] = 'Salary type name is required.'
        elif SalaryType.objects.filter(name__iexact=values['name']).exclude(pk=salary_type.pk).exists():
            errors['name'] = 'Salary type already exists.'

        if not errors:
            salary_type.name = values['name']
            salary_type.description = values['description']
            salary_type.is_active = values['is_active']
            salary_type.save()
            messages.success(request, 'Salary type updated successfully.')
            return redirect('master:salary_list')

    context = {
        'salary_type': salary_type,
        'values': values,
        'errors': errors,
    }
    return render(request, 'master/salary_creation/edit.html', context)


@permission_required('master.delete_salarytype', raise_exception=True)
def salary_delete(request, pk):
    salary_type = get_object_or_404(SalaryType, pk=pk)
    if request.method == 'POST':
        salary_type.delete()
        messages.success(request, 'Salary type deleted successfully.')
        return redirect('master:salary_list')

    context = {
        'salary_type': salary_type,
    }
    return render(request, 'master/salary_creation/confirm_delete.html', context)


@permission_required('master.view_shift', raise_exception=True)
def shift_list(request):
    shifts = Shift.objects.order_by('created_at', 'name')
    search_query = request.GET.get('q', '').strip()
    per_page = request.GET.get('per_page', '').strip() or '10'
    page_number = request.GET.get('page')

    try:
        per_page_value = max(int(per_page), 1)
    except ValueError:
        per_page_value = 10

    if search_query:
        shifts = shifts.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    paginator = Paginator(shifts, per_page_value)
    page_obj = paginator.get_page(page_number)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    base_querystring = query_params.urlencode()
    page_query_base = f'{base_querystring}&' if base_querystring else ''

    context = {
        'shifts': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'search_query': search_query,
        'per_page': per_page_value,
        'base_querystring': base_querystring,
        'page_query_base': page_query_base,
    }
    return render(request, 'master/shift_creation/list.html', context)


@permission_required('master.view_shift', raise_exception=True)
def shift_roster_list(request):
    summary = {
        'total_week': len(WEEK_ROSTERS),
        'total_month': len(MONTH_ROSTERS),
    }
    quick_links = [
        {
            'title': 'Shift Roster Hub',
            'description': 'Choose between week wise or month wise roster planning.',
            'url': reverse('master:shift_roster_main'),
            'cta': 'Open Roster Hub',
        },
        {
            'title': 'Week Wise Roster',
            'description': 'View and manage weekly rosters with quick filters.',
            'url': reverse('master:shift_roster_week'),
            'cta': 'View Week Roster',
        },
        {
            'title': 'Month Wise Roster',
            'description': 'Review month wise schedules and update assignments.',
            'url': reverse('master:shift_roster_month'),
            'cta': 'View Month Roster',
        },
    ]

    context = {
        'summary': summary,
        'quick_links': quick_links,
    }
    return render(request, 'master/shift_roster/list.html', context)


@permission_required('master.view_shift', raise_exception=True)
def shift_roster_main(request):
    context = {
        'week_url': reverse('master:shift_roster_week'),
        'month_url': reverse('master:shift_roster_month'),
    }
    return render(request, 'master/shift_roster/main.html', context)


@permission_required('master.view_shift', raise_exception=True)
def shift_roster_week(request):
    filters = {
        'from_date': request.GET.get('from_date', '').strip(),
        'to_date': request.GET.get('to_date', '').strip(),
        'site': request.GET.get('site', '').strip(),
        'salary_type': request.GET.get('salary_type', '').strip(),
    }

    from_date = _parse_date(filters['from_date'])
    to_date = _parse_date(filters['to_date'])

    rosters = []
    for roster in WEEK_ROSTERS:
        if filters['site'] and filters['site'] != roster['site']:
            continue
        if filters['salary_type'] and filters['salary_type'] != roster['salary_type']:
            continue
        if from_date and roster['entry_date'] < from_date:
            continue
        if to_date and roster['entry_date'] > to_date:
            continue
        rosters.append(roster)

    context = {
        'filters': filters,
        'sites': ROSTER_SITES,
        'salary_types': ROSTER_SALARY_TYPES,
        'rosters': rosters,
    }
    return render(request, 'master/shift_roster/week_roster.html', context)


@permission_required('master.add_shift', raise_exception=True)
def shift_roster_create(request):
    values = {
        'site_name': '',
        'salary_type': '',
        'from_date': '',
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values.update({
            'site_name': request.POST.get('site_name', '').strip(),
            'salary_type': request.POST.get('salary_type', '').strip(),
            'from_date': request.POST.get('from_date', '').strip(),
            'description': request.POST.get('description', '').strip(),
        })

        if not values['site_name']:
            errors['site_name'] = 'Select a site to continue.'
        if not values['salary_type']:
            errors['salary_type'] = 'Choose a salary type.'
        if not values['from_date']:
            errors['from_date'] = 'From date is required.'
        elif not _parse_date(values['from_date']):
            errors['from_date'] = 'Enter a valid date (YYYY-MM-DD).'

        if not errors:
            messages.success(request, 'Week wise roster saved successfully (demo).')
            return redirect('master:shift_roster_week')

    context = {
        'sites': ROSTER_SITES,
        'salary_types': ROSTER_SALARY_TYPES,
        'values': values,
        'errors': errors,
        'week_days': ROSTER_WEEK_DAYS,
        'employees': ROSTER_EMPLOYEES,
    }
    return render(request, 'master/shift_roster/create.html', context)


@permission_required('master.view_shift', raise_exception=True)
def shift_roster_month(request):
    filters = {
        'from_date': request.GET.get('from_date', '').strip(),
        'to_date': request.GET.get('to_date', '').strip(),
        'site': request.GET.get('site', '').strip(),
        'salary_type': request.GET.get('salary_type', '').strip(),
    }

    from_date = _parse_date(filters['from_date'])
    to_date = _parse_date(filters['to_date'])

    rosters = []
    for roster in MONTH_ROSTERS:
        if filters['site'] and filters['site'] != roster['site']:
            continue
        if filters['salary_type'] and filters['salary_type'] != roster['salary_type']:
            continue
        entry_month = roster['entry_month']
        if from_date and entry_month < from_date:
            continue
        if to_date and entry_month > to_date:
            continue
        rosters.append(roster)

    context = {
        'filters': filters,
        'sites': ROSTER_SITES,
        'salary_types': ROSTER_SALARY_TYPES,
        'rosters': rosters,
    }
    return render(request, 'master/shift_roster/month_roster.html', context)


@permission_required('master.change_shift', raise_exception=True)
def shift_roster_month_update(request):
    roster_id = request.GET.get('id')
    selected = None
    if roster_id:
        selected = next((item for item in MONTH_ROSTERS if str(item['id']) == roster_id), None)

    values = {
        'site_name': selected['site'] if selected else '',
        'salary_type': selected['salary_type'] if selected else '',
        'month_date': selected['entry_month'].strftime('%Y-%m-%d') if selected else '',
        'description': '',
    }
    errors = {}

    if request.method == 'POST':
        values.update({
            'site_name': request.POST.get('site_name', '').strip(),
            'salary_type': request.POST.get('salary_type', '').strip(),
            'month_date': request.POST.get('month_date', '').strip(),
            'description': request.POST.get('description', '').strip(),
        })

        if not values['site_name']:
            errors['site_name'] = 'Select a site to continue.'
        if not values['salary_type']:
            errors['salary_type'] = 'Choose a salary type.'
        if not values['month_date']:
            errors['month_date'] = 'Month selection is required.'
        elif not _parse_date(values['month_date']):
            errors['month_date'] = 'Enter a valid month start date.'

        if not errors:
            messages.success(request, 'Month wise roster updated successfully (demo).')
            return redirect('master:shift_roster_month')

    context = {
        'sites': ROSTER_SITES,
        'salary_types': ROSTER_SALARY_TYPES,
        'values': values,
        'errors': errors,
        'week_days': ROSTER_WEEK_DAYS,
        'employees': ROSTER_EMPLOYEES,
    }
    return render(request, 'master/shift_roster/monthUpdate.html', context)


@permission_required('master.add_shift', raise_exception=True)
def shift_edit_form(request):
    shift_id = request.GET.get('id')
    shift = None
    if shift_id:
        shift = get_object_or_404(Shift, pk=shift_id)

    if request.method == 'POST':
        submitted = request.POST.dict()
        name = submitted.get('name', '').strip()
        start_time_raw = submitted.get('start_time', '').strip()
        end_time_raw = submitted.get('end_time', '').strip()
        description = submitted.get('description', '').strip()

        errors = {}
        if not name:
            errors['name'] = 'Shift name is required.'

        def parse_time(value, field):
            if not value:
                errors[field] = 'Time required.'
                return None
            try:
                return datetime.strptime(value, '%H:%M').time()
            except ValueError:
                errors[field] = 'Invalid time format.'
                return None

        start_time = parse_time(start_time_raw, 'start_time')
        end_time = parse_time(end_time_raw, 'end_time')

        if not errors:
            if shift is None:
                Shift.objects.create(
                    name=name,
                    start_time=start_time,
                    end_time=end_time,
                    description=description,
                )
                messages.success(request, 'Shift created successfully.')
            else:
                shift.name = name
                shift.start_time = start_time
                shift.end_time = end_time
                shift.description = description
                shift.save()
                messages.success(request, 'Shift updated successfully.')
            return redirect('master:shift_list')

        context = {
            'errors': errors,
            'values': submitted,
            'shift': shift,
        }
        return render(request, 'master/shift_creation/editForm.html', context)

    initial = {}
    if shift:
        initial = {
            'name': shift.name,
            'start_time': shift.start_time.strftime('%H:%M'),
            'end_time': shift.end_time.strftime('%H:%M'),
            'description': shift.description,
        }

    context = {
        'shift': shift,
        'errors': {},
        'values': initial,
    }
    return render(request, 'master/shift_creation/editForm.html', context)


@permission_required('master.delete_shift', raise_exception=True)
def shift_delete(request, pk):
    shift = get_object_or_404(Shift, pk=pk)
    if request.method == 'POST':
        shift.delete()
        messages.success(request, 'Shift deleted successfully.')
        return redirect('master:shift_list')

    return render(request, 'master/shift_creation/confirm_delete.html', {'shift': shift})


@permission_required('master.view_employee', raise_exception=True)
def employee_list(request):
    staff_status = request.GET.get('staff_status', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    
    # Get all employees from Employee model
    employees = Employee.objects.select_related('company').all()
    
    # Filter by company
    if company_name:
        try:
            employees = employees.filter(company_id=int(company_name))
        except (ValueError, TypeError):
            pass
    
    # Note: Status filter can be added later if needed
    # For now, we'll show all employees
    
    # Get all companies, ordered by billing_name
    companies = Company.objects.all().order_by('billing_name')
    
    context = {
        'employees': employees,
        'companies': companies,
        'staff_status': staff_status,
        'company_name': company_name,
    }
    return render(request, 'master/employee_creation/list.html', context)


@permission_required('master.add_employee', raise_exception=True)
def employee_create(request):
    companies = Company.objects.all().order_by('billing_name')
    last_employee = Employee.objects.order_by('-id').first()
    last_staff_id = getattr(last_employee, 'staff_id', '') if last_employee else ''
    context = {
        'companies': companies,
        'staff': None,
        'last_staff_id': last_staff_id,
    }
    return render(request, 'master/employee_creation/create.html', context)


@permission_required('master.change_employee', raise_exception=True)
def employee_edit(request, pk):
    """Edit employee with all related data."""
    employee = get_object_or_404(
        Employee.objects.select_related('company', 'account_info', 'vehicle_detail'),
        pk=pk
    )
    companies = Company.objects.all().order_by('billing_name')
    
    # Get related data
    dependents = employee.dependents.all()
    qualifications = employee.qualifications.all()
    experiences = employee.experiences.all()
    assets = employee.asset_assignments.all()
    
    context = {
        'employee': employee,
        'companies': companies,
        'dependents': dependents,
        'qualifications': qualifications,
        'experiences': experiences,
        'assets': assets,
        'staff': employee,  # For template compatibility
    }
    return render(request, 'master/employee_creation/edit.html', context)


@permission_required('master.delete_employee', raise_exception=True)
@require_POST
def employee_delete(request, pk):
    """Delete employee and all related data."""
    employee = get_object_or_404(Employee, pk=pk)
    employee_name = employee.staff_name
    
    try:
        # Try to delete the employee
        employee.delete()
        messages.success(request, f'Employee {employee_name} deleted successfully.')
        return redirect('master:employee_list')
        
    except ProtectedError as e:
        # Handle protected relationships that prevent deletion
        protected_objects = list(e.protected_objects)
        
        # Check if the protected objects are CompOffEntry records
        if protected_objects and hasattr(protected_objects[0], 'work_date'):
            # CompOffEntry records
            entry_dates = [str(obj.work_date) for obj in protected_objects[:5]]  # Show first 5
            entry_count = len(protected_objects)
            
            if entry_count > 5:
                error_msg = (
                    f'Cannot delete employee {employee_name} because they have {entry_count} '
                    f'Comp-Off Entry record(s) associated. Please delete the Comp-Off entries first. '
                    f'Entries include: {", ".join(entry_dates)}...'
                )
            else:
                error_msg = (
                    f'Cannot delete employee {employee_name} because they have {entry_count} '
                    f'Comp-Off Entry record(s) associated: {", ".join(entry_dates)}. '
                    f'Please delete the Comp-Off entries first.'
                )
        else:
            # Other protected relationships
            error_msg = (
                f'Cannot delete employee {employee_name} because they are referenced by '
                f'{len(protected_objects)} other record(s). Please remove the related records first.'
            )
        
        messages.error(request, error_msg)
        return redirect('master:employee_list')


@permission_required('master.view_employee', raise_exception=True)
def employee_export_excel(request):
    """Export employee list to Excel with filters."""
    staff_status = request.GET.get('staff_status', '').strip()
    company_name = request.GET.get('company_name', '').strip()
    
    # Get all employees from Employee model
    employees = Employee.objects.select_related('company').all()
    
    # Filter by company
    if company_name:
        try:
            employees = employees.filter(company_id=int(company_name))
        except (ValueError, TypeError):
            pass
    
    # Filter by status (if implemented in future)
    # if staff_status and staff_status != '0':
    #     employees = employees.filter(...)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee List"
    
    # Header row styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define headers
    headers = [
        'S.No', 'Staff ID', 'Staff Name', 'DOB', 'Gender', 
        'Designation', 'Department', 'Work Location', 
        'Company', 'Date of Join', 'Contact', 'Email'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, employee in enumerate(employees, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)  # S.No
        ws.cell(row=row_num, column=2, value=employee.staff_id or '')
        ws.cell(row=row_num, column=3, value=employee.staff_name or '')
        ws.cell(row=row_num, column=4, value=employee.date_of_birth.strftime('%d-%m-%Y') if employee.date_of_birth else '')
        ws.cell(row=row_num, column=5, value=employee.gender or '')
        ws.cell(row=row_num, column=6, value=employee.designation or '')
        ws.cell(row=row_num, column=7, value=employee.department or '')
        ws.cell(row=row_num, column=8, value=employee.work_location or '')
        ws.cell(row=row_num, column=9, value=employee.company.billing_name if employee.company else '')
        ws.cell(row=row_num, column=10, value=employee.date_of_join.strftime('%d-%m-%Y') if employee.date_of_join else '')
        ws.cell(row=row_num, column=11, value=employee.personal_contact or employee.office_contact or '')
        ws.cell(row=row_num, column=12, value=employee.personal_email or employee.office_email or '')
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        # Check header length
        max_length = max(max_length, len(str(header)))
        # Check data length
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        # Set column width with some padding
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set row height for header
    ws.row_dimensions[1].height = 25
    
    # Create HTTP response with Excel content
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'employee_list_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response


def _clean_required(data, field, label, errors):
    value = data.get(field, '').strip()
    if not value:
        errors[field] = f'{label} is required.'
    return value


def _clean_text_only(data, field, label, errors):
    """Validate text fields that should only contain letters, spaces, dots, hyphens, and apostrophes."""
    value = _clean_required(data, field, label, errors)
    if value:
        import re
        # Allow letters, spaces, dots, hyphens, apostrophes (for names like O'Brien, Mary-Jane, etc.)
        if not re.match(r'^[a-zA-Z\s\.\-\']+$', value):
            errors[field] = f'{label} should only contain letters, spaces, dots, hyphens, and apostrophes.'
    return value


def _clean_optional(data, field):
    return data.get(field, '').strip()


def _clean_phone(data, field, label, errors, length=10):
    value = _clean_optional(data, field)
    if value:
        # Remove spaces and hyphens
        cleaned = value.replace(' ', '').replace('-', '')
        if not cleaned.isdigit() or len(cleaned) != length:
            errors[field] = f'{label} must be exactly {length} digits.'
    return value


def _clean_aadhar(data, field, label, errors):
    """Validate Aadhar number (12 digits)."""
    value = _clean_optional(data, field)
    if value:
        cleaned = value.replace(' ', '').replace('-', '')
        if not cleaned.isdigit() or len(cleaned) != 12:
            errors[field] = f'{label} must be exactly 12 digits.'
        return cleaned
    return value


def _clean_pan(data, field, label, errors):
    """Validate PAN number (5 letters + 4 digits + 1 letter)."""
    value = _clean_optional(data, field)
    if value:
        cleaned = value.replace(' ', '').upper()
        import re
        if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]{1}$', cleaned):
            errors[field] = f'{label} must be in format: ABCDE1234F (5 letters, 4 digits, 1 letter).'
        return cleaned
    return value


def _clean_pincode(data, field, label, errors):
    """Validate pincode (6 digits)."""
    value = _clean_optional(data, field)
    if value:
        if not value.isdigit() or len(value) != 6:
            errors[field] = f'{label} must be exactly 6 digits.'
    return value


def _clean_ifsc(data, field, label, errors):
    """Validate IFSC code (4 letters + 0 + 6 alphanumeric)."""
    value = _clean_optional(data, field)
    if value:
        cleaned = value.replace(' ', '').upper()
        import re
        if not re.match(r'^[A-Z]{4}0[A-Z0-9]{6}$', cleaned):
            errors[field] = f'{label} must be in format: ABCD0123456 (4 letters, 0, 6 alphanumeric).'
        return cleaned
    return value


def _clean_account_number(data, field, label, errors):
    """Validate account number (9-18 digits)."""
    value = _clean_optional(data, field)
    if value:
        cleaned = value.replace(' ', '').replace(',', '')
        if not cleaned.isdigit() or len(cleaned) < 9 or len(cleaned) > 18:
            errors[field] = f'{label} must be 9-18 digits.'
        return cleaned
    return value


def _clean_percentage(data, field, label, errors):
    """Validate percentage (0-100)."""
    value = _clean_optional(data, field)
    if value:
        try:
            num = float(value)
            if num < 0 or num > 100:
                errors[field] = f'{label} must be between 0 and 100.'
        except ValueError:
            errors[field] = f'{label} must be a valid number.'
    return value


def _clean_year(data, field, label, errors):
    """Validate year (1900 to current year)."""
    value = _clean_optional(data, field)
    if value:
        try:
            year = int(value)
            from datetime import datetime
            current_year = datetime.now().year
            if year < 1900 or year > current_year:
                errors[field] = f'{label} must be between 1900 and {current_year}.'
        except ValueError:
            errors[field] = f'{label} must be a valid year.'
    return value


def _clean_salary(data, field, label, errors):
    """Validate salary/amount (positive number)."""
    value = _clean_optional(data, field)
    if value:
        cleaned = value.replace(',', '').replace(' ', '')
        try:
            num = float(cleaned)
            if num < 0:
                errors[field] = f'{label} must be a positive number.'
        except ValueError:
            errors[field] = f'{label} must be a valid number.'
        return cleaned
    return value


def _clean_vehicle_reg(data, field, label, errors):
    """Validate vehicle registration number (Indian format)."""
    value = _clean_optional(data, field)
    if value:
        import re
        cleaned = value.replace(' ', '').upper()
        if not re.match(r'^[A-Z]{2}[-\s]?[0-9]{1,2}[-\s]?[A-Z]{1,2}[-\s]?[0-9]{4}$', cleaned):
            errors[field] = f'{label} must be in format: TN-09-AB-1234.'
        return cleaned
    return value


def _clean_license(data, field, label, errors):
    """Validate license number (10-15 alphanumeric)."""
    value = _clean_optional(data, field)
    if value:
        import re
        cleaned = value.replace(' ', '').upper()
        if not re.match(r'^[A-Z0-9]{10,15}$', cleaned):
            errors[field] = f'{label} must be 10-15 alphanumeric characters.'
        return cleaned
    return value


def _clean_date_not_future(data, field, label, errors):
    """Validate date is not in the future."""
    value = _clean_optional(data, field)
    if value:
        from datetime import datetime
        try:
            date_value = datetime.strptime(value, '%Y-%m-%d').date()
            today = datetime.now().date()
            if date_value > today:
                errors[field] = f'{label} cannot be a future date.'
            # Check minimum age for DOB (18 years)
            if 'birth' in field.lower() or 'dob' in field.lower():
                from datetime import date
                age = today.year - date_value.year - ((today.month, today.day) < (date_value.month, date_value.day))
                if age < 18:
                    errors[field] = f'{label} must be at least 18 years ago.'
        except ValueError:
            errors[field] = f'{label} must be a valid date.'
    return value


# Indian States and Union Territories
INDIAN_STATES = [
    'Andhra Pradesh', 'Arunachal Pradesh', 'Assam', 'Bihar', 'Chhattisgarh',
    'Goa', 'Gujarat', 'Haryana', 'Himachal Pradesh', 'Jharkhand', 'Karnataka',
    'Kerala', 'Madhya Pradesh', 'Maharashtra', 'Manipur', 'Meghalaya', 'Mizoram',
    'Nagaland', 'Odisha', 'Punjab', 'Rajasthan', 'Sikkim', 'Tamil Nadu', 'Telangana',
    'Tripura', 'Uttar Pradesh', 'Uttarakhand', 'West Bengal',
    'Andaman and Nicobar Islands', 'Chandigarh', 'Dadra and Nagar Haveli and Daman and Diu',
    'Delhi', 'Jammu and Kashmir', 'Ladakh', 'Lakshadweep', 'Puducherry'
]

# Major Indian Cities by State
INDIAN_CITIES_BY_STATE = {
    'Andhra Pradesh': ['Visakhapatnam', 'Vijayawada', 'Guntur', 'Nellore', 'Kurnool', 'Rajahmundry', 'Tirupati', 'Kakinada'],
    'Arunachal Pradesh': ['Itanagar', 'Naharlagun', 'Tawang', 'Pasighat'],
    'Assam': ['Guwahati', 'Silchar', 'Dibrugarh', 'Jorhat', 'Nagaon', 'Tinsukia'],
    'Bihar': ['Patna', 'Gaya', 'Bhagalpur', 'Muzaffarpur', 'Purnia', 'Darbhanga'],
    'Chhattisgarh': ['Raipur', 'Bhilai', 'Bilaspur', 'Korba', 'Durg', 'Rajpur'],
    'Goa': ['Panaji', 'Vasco da Gama', 'Margao', 'Mapusa'],
    'Gujarat': ['Ahmedabad', 'Surat', 'Vadodara', 'Rajkot', 'Bhavnagar', 'Jamnagar', 'Gandhinagar'],
    'Haryana': ['Gurgaon', 'Faridabad', 'Panipat', 'Ambala', 'Yamunanagar', 'Karnal', 'Hisar'],
    'Himachal Pradesh': ['Shimla', 'Mandi', 'Solan', 'Dharamshala', 'Kullu'],
    'Jharkhand': ['Ranchi', 'Jamshedpur', 'Dhanbad', 'Bokaro', 'Hazaribagh'],
    'Karnataka': ['Bangalore', 'Mysore', 'Hubli', 'Mangalore', 'Belgaum', 'Gulbarga', 'Davangere'],
    'Kerala': ['Kochi', 'Thiruvananthapuram', 'Kozhikode', 'Thrissur', 'Kollam', 'Alappuzha'],
    'Madhya Pradesh': ['Bhopal', 'Indore', 'Gwalior', 'Jabalpur', 'Ujjain', 'Raipur'],
    'Maharashtra': ['Mumbai', 'Pune', 'Nagpur', 'Nashik', 'Aurangabad', 'Solapur', 'Thane'],
    'Manipur': ['Imphal', 'Thoubal', 'Bishnupur'],
    'Meghalaya': ['Shillong', 'Tura', 'Jowai'],
    'Mizoram': ['Aizawl', 'Lunglei', 'Champhai'],
    'Nagaland': ['Kohima', 'Dimapur', 'Mokokchung'],
    'Odisha': ['Bhubaneswar', 'Cuttack', 'Rourkela', 'Berhampur', 'Sambalpur'],
    'Punjab': ['Ludhiana', 'Amritsar', 'Jalandhar', 'Patiala', 'Bathinda', 'Mohali'],
    'Rajasthan': ['Jaipur', 'Jodhpur', 'Kota', 'Bikaner', 'Ajmer', 'Udaipur', 'Bhilwara'],
    'Sikkim': ['Gangtok', 'Namchi', 'Mangan'],
    'Tamil Nadu': ['Chennai', 'Coimbatore', 'Madurai', 'Tiruchirappalli', 'Salem', 'Tirunelveli', 'Erode'],
    'Telangana': ['Hyderabad', 'Warangal', 'Nizamabad', 'Karimnagar', 'Khammam'],
    'Tripura': ['Agartala', 'Udaipur', 'Dharmanagar'],
    'Uttar Pradesh': ['Lucknow', 'Kanpur', 'Agra', 'Varanasi', 'Allahabad', 'Meerut', 'Ghaziabad', 'Noida'],
    'Uttarakhand': ['Dehradun', 'Haridwar', 'Roorkee', 'Haldwani', 'Rudrapur'],
    'West Bengal': ['Kolkata', 'Howrah', 'Durgapur', 'Asansol', 'Siliguri', 'Bardhaman'],
    'Andaman and Nicobar Islands': ['Port Blair'],
    'Chandigarh': ['Chandigarh'],
    'Dadra and Nagar Haveli and Daman and Diu': ['Daman', 'Diu', 'Silvassa'],
    'Delhi': ['New Delhi', 'Delhi'],
    'Jammu and Kashmir': ['Srinagar', 'Jammu', 'Anantnag', 'Baramulla'],
    'Ladakh': ['Leh', 'Kargil'],
    'Lakshadweep': ['Kavaratti'],
    'Puducherry': ['Puducherry', 'Karaikal', 'Mahe', 'Yanam']
}


@login_required
@require_http_methods(["GET"])
def get_countries(request):
    """API endpoint to get list of countries."""
    # Fallback countries list
    fallback_countries = [
        'India', 'United States', 'United Kingdom', 'Canada', 'Australia', 
        'Germany', 'France', 'Japan', 'China', 'Singapore', 'United Arab Emirates',
        'Saudi Arabia', 'Qatar', 'Kuwait', 'Oman', 'Bahrain', 'Malaysia',
        'Thailand', 'Indonesia', 'Philippines', 'South Korea', 'New Zealand',
        'South Africa', 'Brazil', 'Mexico', 'Argentina', 'Chile', 'Russia'
    ]
    
    try:
        # Try to use REST Countries API if requests library is available
        try:
            import requests  # type: ignore[reportMissingModuleSource]
            response = requests.get('https://restcountries.com/v3.1/all?fields=name', timeout=5)
            if response.status_code == 200:
                countries = sorted([country['name']['common'] for country in response.json()])
                return JsonResponse({'success': True, 'countries': countries})
        except ImportError:
            # requests library not installed, use fallback
            pass
        except Exception:
            # API call failed, use fallback
            pass
    except Exception:
        # Any other error, use fallback
        pass
    
    # Return fallback countries
    return JsonResponse({'success': True, 'countries': sorted(fallback_countries)})


@login_required
@require_http_methods(["GET"])
def get_states(request):
    """API endpoint to get list of states (Indian states by default, or by country)."""
    country = request.GET.get('country', 'India')
    
    if country == 'India':
        states = sorted(INDIAN_STATES)
    else:
        # For other countries, return empty or use an API
        # You can integrate with a states API here
        states = []
    
    return JsonResponse({'success': True, 'states': states})


@login_required
@require_http_methods(["GET"])
def get_cities(request):
    """API endpoint to get list of cities by state."""
    state = request.GET.get('state', '')
    country = request.GET.get('country', 'India')
    
    if country == 'India' and state:
        cities = INDIAN_CITIES_BY_STATE.get(state, [])
        cities = sorted(cities)
    else:
        cities = []
    
    return JsonResponse({'success': True, 'cities': cities})


def _clean_email(data, field, label, errors):
    value = _clean_optional(data, field)
    if value:
        try:
            validate_email(value)
        except ValidationError:
            errors[field] = f'Enter a valid {label}.'
    return value


def _clean_aadhar(data, field, errors):
    value = _clean_optional(data, field).replace(' ', '')
    if value:
        if not value.isdigit() or len(value) != 12:
            errors[field] = 'Aadhar number must be 12 digits.'
    return value


def _clean_date(data, field, label, errors):
    value = data.get(field, '').strip()
    if not value:
        errors[field] = f'{label} is required.'
        return None
    parsed = _parse_date(value)
    if not parsed:
        errors[field] = f'Enter a valid date for {label}.'
    return parsed


def _validate_staff_details(data, unique_id):
    """
    Validate Staff Details tab - all fields from Staff Details navigation page.
    Returns (errors dict, validated_data dict).
    """
    errors = {}
    validated_data = {}
    
    # Personal Details
    staff_name = _clean_text_only(data, 'staff_name', 'Staff Name', errors)
    staff_id = _clean_required(data, 'staff_id', 'Staff ID', errors)
    
    if staff_id:
        # Check if staff_id already exists (excluding current employee if editing)
        existing = Employee.objects.filter(staff_id=staff_id)
        if unique_id:
            existing = existing.exclude(unique_id=unique_id)
        if existing.exists():
            errors['staff_id'] = 'Staff ID already exists.'
    
    gender = _clean_required(data, 'gender', 'Gender', errors)
    if gender and gender not in [choice[0] for choice in Employee.GENDER_CHOICES]:
        errors['gender'] = 'Invalid gender selected.'
    
    father_name = _clean_text_only(data, 'father_name', 'Father Name', errors)
    
    date_of_birth = _clean_date(data, 'date_of_birth', 'Date of Birth', errors)
    doc_dob = _clean_date(data, 'doc_dob', 'Document DOB', errors)
    
    age = None
    age_str = data.get('age', '').strip()
    if not age_str:
        errors['age'] = 'Age is required.'
    else:
        try:
            age = int(age_str)
            if age < 18 or age > 70:
                errors['age'] = 'Age must be between 18 and 70.'
        except ValueError:
            errors['age'] = 'Age must be a valid number.'
    
    marital_status = _clean_required(data, 'marital_status', 'Marital Status', errors)
    if marital_status and marital_status not in [choice[0] for choice in Employee.MARITAL_CHOICES]:
        errors['marital_status'] = 'Invalid marital status.'
    
    personal_contact = _clean_phone(data, 'personal_contact', 'Personal Contact', errors)
    if not personal_contact:
        errors['personal_contact'] = 'Personal Contact number is required.'
    
    office_contact = _clean_phone(data, 'office_contact', 'Office Contact', errors)
    if not office_contact:
        errors['office_contact'] = 'Office Contact number is required.'
    
    personal_email = _clean_email(data, 'personal_email', 'Personal Email', errors)
    if not personal_email:
        errors['personal_email'] = 'Personal Email is required.'
    
    office_email = _clean_email(data, 'office_email', 'Office Email', errors)
    if not office_email:
        errors['office_email'] = 'Office Email is required.'
    
    aadhar_no = _clean_aadhar(data, 'aadhar_no', 'Aadhar No', errors)
    if not aadhar_no:
        errors['aadhar_no'] = 'Aadhar number is required.'
    
    pan_no = _clean_pan(data, 'pan_no', 'PAN No', errors)
    if not pan_no:
        errors['pan_no'] = 'PAN number is required.'
    
    medical_claim_str = _clean_required(data, 'medical_claim', 'Medical Claim', errors)
    if medical_claim_str and medical_claim_str not in ['Yes', 'No']:
        errors['medical_claim'] = 'Invalid medical claim value.'
    medical_claim = medical_claim_str == 'Yes' if medical_claim_str else False
    
    blood_group = _clean_required(data, 'blood_group', 'Blood Group', errors)
    qualification = _clean_required(data, 'qualification', 'Qualification', errors)
    
    # Address fields - all required
    present_country = _clean_required(data, 'pre_country', 'Present Country', errors)
    present_state = _clean_required(data, 'pre_state', 'Present State', errors)
    present_city = _clean_required(data, 'pre_city', 'Present City', errors)
    present_building = _clean_required(data, 'pre_building', 'Present Building No', errors)
    present_street = _clean_required(data, 'pre_street', 'Present Street', errors)
    present_area = _clean_required(data, 'pre_area', 'Present Area', errors)
    present_pincode = _clean_pincode(data, 'pre_pincode', 'Present Pincode', errors)
    if not present_pincode:
        errors['pre_pincode'] = 'Present Pincode is required.'
    
    permanent_country = _clean_required(data, 'perm_country', 'Permanent Country', errors)
    permanent_state = _clean_required(data, 'perm_state', 'Permanent State', errors)
    permanent_city = _clean_required(data, 'perm_city', 'Permanent City', errors)
    permanent_building = _clean_required(data, 'perm_building', 'Permanent Building No', errors)
    permanent_street = _clean_required(data, 'perm_street', 'Permanent Street', errors)
    permanent_area = _clean_required(data, 'perm_area', 'Permanent Area', errors)
    permanent_pincode = _clean_pincode(data, 'perm_pincode', 'Permanent Pincode', errors)
    if not permanent_pincode:
        errors['perm_pincode'] = 'Permanent Pincode is required.'
    
    # Official Details
    date_of_join = _clean_date(data, 'date_of_join', 'Date of Join', errors)
    designation = _clean_required(data, 'designation', 'Designation', errors)
    department = _clean_required(data, 'department', 'Department', errors)
    work_location = _clean_required(data, 'work_location', 'Work Location', errors)
    esi_no = _clean_required(data, 'esi_no', 'ESI No', errors)
    pf_no = _clean_required(data, 'pf_no', 'PF No', errors)
    biometric_id = _clean_required(data, 'biometric_id', 'Bio Metric ID', errors)
    
    company_id = data.get('company_name', '').strip()
    company = None
    if not company_id:
        errors['company_name'] = 'Company selection is required.'
    else:
        try:
            company = Company.objects.get(pk=int(company_id))
        except (Company.DoesNotExist, ValueError, TypeError):
            errors['company_name'] = 'Invalid company selected.'
    
    salary_category = _clean_required(data, 'salary_category', 'Salary Category', errors)
    if salary_category and salary_category not in [choice[0] for choice in Employee.SALARY_CATEGORY_CHOICES]:
        errors['salary_category'] = 'Invalid salary category.'
    
    premises_type = data.get('premises_type', Employee.PREMISES_OUT).strip()
    if premises_type not in [choice[0] for choice in Employee.PREMISES_CHOICES]:
        premises_type = Employee.PREMISES_OUT
    
    branch = _clean_required(data, 'branch_ids', 'Branch', errors)
    attendance_setting = _clean_required(data, 'attendance_setting', 'Attendance Setting', errors)
    reporting_officer = _clean_required(data, 'reporting_officer', 'Reporting Officer', errors)
    
    validated_data = {
        'staff_name': staff_name,
        'staff_id': staff_id,
        'gender': gender,
        'father_name': father_name,
        'date_of_birth': date_of_birth,
        'document_date_of_birth': doc_dob,
        'age': age,
        'marital_status': marital_status,
        'personal_contact': personal_contact,
        'office_contact': office_contact,
        'personal_email': personal_email,
        'office_email': office_email,
        'aadhar_no': aadhar_no,
        'pan_no': pan_no,
        'medical_claim': medical_claim,
        'blood_group': blood_group,
        'qualification': qualification,
        'present_country': present_country,
        'present_state': present_state,
        'present_city': present_city,
        'present_building': present_building,
        'present_street': present_street,
        'present_area': present_area,
        'present_pincode': present_pincode,
        'permanent_country': permanent_country,
        'permanent_state': permanent_state,
        'permanent_city': permanent_city,
        'permanent_building': permanent_building,
        'permanent_street': permanent_street,
        'permanent_area': permanent_area,
        'permanent_pincode': permanent_pincode,
        'date_of_join': date_of_join,
        'designation': designation,
        'department': department,
        'work_location': work_location,
        'esi_no': esi_no,
        'pf_no': pf_no,
        'biometric_id': biometric_id,
        'company': company,
        'salary_category': salary_category,
        'premises_type': premises_type,
        'branch': branch,
        'attendance_setting': attendance_setting,
        'reporting_officer': reporting_officer,
    }
    
    return errors, validated_data


def _validate_dependent_details(data):
    """
    Validate Dependent Details tab - all fields from Dependent Details navigation page.
    All fields are required.
    Returns (errors dict, validated_data dict).
    """
    errors = {}
    validated_data = {}
    
    # All fields are required
    rel_relationship = _clean_required(data, 'relationship', 'Relationship', errors)
    rel_name = _clean_required(data, 'rel_name', 'Dependent Name', errors)
    
    # Validate gender (required)
    rel_gender = _clean_required(data, 'rel_gender', 'Gender', errors)
    if rel_gender and rel_gender not in [choice[0] for choice in Employee.GENDER_CHOICES]:
        errors['rel_gender'] = 'Invalid gender selected.'
    
    # Validate date of birth (required)
    rel_date_of_birth = _clean_date(data, 'rel_date_of_birth', 'Date of Birth', errors)
    
    # Validate Aadhar number (required)
    rel_aadhar_no = _clean_aadhar(data, 'rel_aadhar_no', 'Aadhar No', errors)
    if not rel_aadhar_no:
        errors['rel_aadhar_no'] = 'Aadhar number is required.'
    
    # All other fields are required
    occupation = _clean_text_only(data, 'occupation', 'Occupation', errors)
    standard = _clean_required(data, 'standard', 'Standard', errors)
    school = _clean_text_only(data, 'school', 'School', errors)
    existing_illness = _clean_required(data, 'existing_illness', 'Existing Illness', errors)
    description = _clean_required(data, 'description', 'Description', errors)
    existing_insurance = _clean_required(data, 'existing_insurance', 'Existing Insurance', errors)
    insurance_no = _clean_required(data, 'insurance_no', 'Insurance No', errors)
    physically_challenged = _clean_required(data, 'physically_challenged', 'Physically Challenged', errors)
    remarks = _clean_required(data, 'remarks', 'Remarks', errors)
    
    validated_data = {
        'rel_relationship': rel_relationship,
        'rel_name': rel_name,
        'rel_gender': rel_gender,
        'rel_date_of_birth': rel_date_of_birth,
        'rel_aadhar_no': rel_aadhar_no,
        'occupation': occupation,
        'standard': standard,
        'school': school,
        'existing_illness': existing_illness,
        'description': description,
        'existing_insurance': existing_insurance,
        'insurance_no': insurance_no,
        'physically_challenged': physically_challenged,
        'remarks': remarks,
    }
    
    return errors, validated_data


def _validate_account_details(data):
    """
    Validate Account Details tab - all fields from Account Details navigation page.
    All fields are required.
    Returns (errors dict, validated_data dict).
    """
    errors = {}
    validated_data = {}
    
    # All fields are required
    bank_status = _clean_required(data, 'bank_status', 'Bank Status', errors)
    if bank_status and bank_status not in [choice[0] for choice in EmployeeAccountInfo.BANK_STATUS_CHOICES]:
        errors['bank_status'] = 'Invalid bank status selected.'
    
    salary_type = _clean_required(data, 'salary_type', 'Salary Type', errors)
    if salary_type and salary_type not in [choice[0] for choice in EmployeeAccountInfo.SALARY_TYPE_CHOICES]:
        errors['salary_type'] = 'Invalid salary type selected.'
    
    accountant_name = _clean_text_only(data, 'accountant_name', 'Accountant Name', errors)
    
    account_no = _clean_account_number(data, 'account_no', 'Account Number', errors)
    if not account_no:
        errors['account_no'] = 'Account number is required.'
    
    bank_name = _clean_text_only(data, 'bank_name', 'Bank Name', errors)
    
    ifsc_code = _clean_ifsc(data, 'ifsc_code', 'IFSC Code', errors)
    if not ifsc_code:
        errors['ifsc_code'] = 'IFSC code is required.'
    
    bank_contact_no = _clean_phone(data, 'bank_contact_no', 'Bank Contact', errors)
    if not bank_contact_no:
        errors['bank_contact_no'] = 'Bank Contact number is required.'
    
    bank_address = _clean_required(data, 'bank_address', 'Bank Address', errors)
    
    validated_data = {
        'bank_status': bank_status,
        'salary_type': salary_type,
        'accountant_name': accountant_name,
        'account_no': account_no,
        'bank_name': bank_name,
        'ifsc_code': ifsc_code,
        'bank_contact_no': bank_contact_no,
        'bank_address': bank_address,
    }
    
    return errors, validated_data


def _validate_qualification_details(data):
    """
    Validate Qualification Details tab - all fields from Qualification Details navigation page.
    All fields are required.
    Returns (errors dict, validated_data dict).
    """
    errors = {}
    validated_data = {}
    
    # All fields are required
    education_type = _clean_required(data, 'education_type', 'Education Type', errors)
    degree = _clean_required(data, 'degree', 'Degree', errors)
    college_name = _clean_text_only(data, 'college_name', 'College Name', errors)
    
    year_passing = _clean_year(data, 'year_passing', 'Year of Passing', errors)
    if not year_passing:
        errors['year_passing'] = 'Year of Passing is required.'
    
    percentage = _clean_percentage(data, 'percentage', 'Percentage', errors)
    if not percentage:
        errors['percentage'] = 'Percentage is required.'
    
    university = _clean_text_only(data, 'university', 'University', errors)
    
    # Note: File validation will be done in the save function where request.FILES is available
    validated_data = {
        'education_type': education_type,
        'degree': degree,
        'college_name': college_name,
        'year_passing': year_passing,
        'percentage': percentage,
        'university': university,
    }
    
    return errors, validated_data


def _validate_experience_details(data):
    """
    Validate Experience Details tab - all fields from Experience Details navigation page.
    All fields are required.
    Returns (errors dict, validated_data dict).
    """
    errors = {}
    validated_data = {}
    
    # All fields are required
    exp_company_name = _clean_text_only(data, 'exp_company_name', 'Company', errors)
    designation_name = _clean_text_only(data, 'designation_name', 'Designation', errors)
    salary_amt = _clean_salary(data, 'salary_amt', 'Salary', errors)
    if not salary_amt:
        errors['salary_amt'] = 'Salary is required.'
    
    join_month = _clean_required(data, 'join_month', 'Joining Month', errors)
    if join_month:
        try:
            join_date = datetime.strptime(join_month + '-01', '%Y-%m-%d')
        except ValueError:
            errors['join_month'] = 'Enter a valid month for joining month.'
    
    relieve_month = _clean_required(data, 'relieve_month', 'Relieving Month', errors)
    if relieve_month:
        try:
            relieve_date = datetime.strptime(relieve_month + '-01', '%Y-%m-%d')
        except ValueError:
            errors['relieve_month'] = 'Enter a valid month for relieving month.'
    
    # Validate that relieving month is after joining month
    if join_month and relieve_month and not errors.get('join_month') and not errors.get('relieve_month'):
        try:
            join_date = datetime.strptime(join_month + '-01', '%Y-%m-%d')
            relieve_date = datetime.strptime(relieve_month + '-01', '%Y-%m-%d')
            if relieve_date < join_date:
                errors['relieve_month'] = 'Relieving month must be after joining month.'
        except ValueError:
            pass  # Already handled above
    
    exp = _clean_required(data, 'exp', 'Experience (years)', errors)
    if exp:
        try:
            exp_float = float(exp)
            if exp_float < 0:
                errors['exp'] = 'Experience must be a positive number.'
        except ValueError:
            errors['exp'] = 'Experience must be a valid number.'
    
    # Note: File validation will be done in the save function where request.FILES is available
    validated_data = {
        'exp_company_name': exp_company_name,
        'designation_name': designation_name,
        'salary_amt': salary_amt,
        'join_month': join_month,
        'relieve_month': relieve_month,
        'exp': exp,
    }
    
    return errors, validated_data


def _validate_asset_vehicle_details(data):
    """
    Validate Asset/Vehicle Details tab - all fields from Asset/Vehicle Details navigation page.
    All fields are required.
    Returns (errors dict, validated_data dict).
    """
    errors = {}
    validated_data = {}
    
    # Asset Details - all required
    asset_name = _clean_required(data, 'asset_name', 'Asset Name', errors)
    item_no = _clean_required(data, 'item_no', 'Serial / Item No', errors)
    
    qty = _clean_required(data, 'qty', 'Quantity', errors)
    if qty:
        try:
            qty_int = int(qty)
            if qty_int < 1:
                errors['qty'] = 'Quantity must be at least 1.'
        except ValueError:
            errors['qty'] = 'Quantity must be a valid number.'
    
    status = _clean_required(data, 'status', 'Status', errors)
    if status and status not in [choice[0] for choice in EmployeeAssetAssignment.STATUS_CHOICES]:
        errors['status'] = 'Invalid asset status selected.'
    
    # License Details - all required
    license_mode = _clean_required(data, 'license_mode', 'License Mode', errors)
    if license_mode and license_mode not in [choice[0] for choice in EmployeeAssetAssignment.LICENSE_MODE_CHOICES]:
        errors['license_mode'] = 'Invalid license mode selected.'
    
    dri_license_no = _clean_required(data, 'dri_license_no', 'License No', errors)
    
    valid_from = _clean_date(data, 'valid_from', 'License Validity From', errors)
    valid_to = _clean_date(data, 'valid_to', 'License Validity To', errors)
    if valid_from and valid_to and valid_from > valid_to:
        errors['valid_to'] = 'License validity end date must be after start date.'
    
    # Vehicle Details - all required
    veh_reg_no = _clean_required(data, 'veh_reg_no', 'Vehicle Reg. No', errors)
    vehicle_type = _clean_required(data, 'vehicle_type', 'Vehicle Type', errors)
    vehicle_company = _clean_text_only(data, 'vehicle_company', 'Vehicle Company', errors)
    vehicle_owner = _clean_text_only(data, 'vehicle_owner', 'Vehicle Owner', errors)
    
    registration_year = _clean_date(data, 'registration_year', 'Year of Registration', errors)
    
    rc_no = _clean_required(data, 'rc_no', 'RC No', errors)
    rc_validity_from = _clean_date(data, 'rc_validity_from', 'RC Validity From', errors)
    rc_validity_to = _clean_date(data, 'rc_validity_to', 'RC Validity To', errors)
    if rc_validity_from and rc_validity_to and rc_validity_from > rc_validity_to:
        errors['rc_validity_to'] = 'RC validity end date must be after start date.'
    
    ins_no = _clean_required(data, 'ins_no', 'Insurance No', errors)
    ins_validity_from = _clean_date(data, 'ins_validity_from', 'Insurance Validity From', errors)
    ins_validity_to = _clean_date(data, 'ins_validity_to', 'Insurance Validity To', errors)
    if ins_validity_from and ins_validity_to and ins_validity_from > ins_validity_to:
        errors['ins_validity_to'] = 'Insurance validity end date must be after start date.'
    
    validated_data = {
        'asset_name': asset_name,
        'item_no': item_no,
        'qty': int(qty) if qty and qty.isdigit() else 1,
        'status': status or EmployeeAssetAssignment.STATUS_ISSUED,
        'veh_reg_no': veh_reg_no,
        'license_mode': license_mode,
        'dri_license_no': dri_license_no,
        'valid_from': valid_from,
        'valid_to': valid_to,
        'vehicle_type': vehicle_type,
        'vehicle_company': vehicle_company,
        'vehicle_owner': vehicle_owner,
        'registration_year': registration_year,
        'rc_no': rc_no,
        'rc_validity_from': rc_validity_from,
        'rc_validity_to': rc_validity_to,
        'ins_no': ins_no,
        'ins_validity_from': ins_validity_from,
        'ins_validity_to': ins_validity_to,
    }
    
    return errors, validated_data


@permission_required('master.add_employee', raise_exception=True)
@require_POST
def employee_staff_save(request):
    """Save employee data from all tabs with comprehensive validation."""
    data = request.POST
    
    # Generate unique_id if not provided
    unique_id = data.get('unique_id', '').strip()
    if not unique_id:
        unique_id = str(uuid.uuid4())
    
    # Check if this is final save (has final_save flag) or intermediate save
    is_final_save = data.get('final_save', 'false').lower() == 'true'
    
    # Check if arrays are present (multiple records format)
    has_dependents_array = any(key.startswith('dependents[') for key in data.keys())
    has_accounts_array = any(key.startswith('accounts[') for key in data.keys())
    has_qualifications_array = any(key.startswith('qualifications[') for key in data.keys())
    has_experiences_array = any(key.startswith('experiences[') for key in data.keys())
    has_assets_array = any(key.startswith('assets[') for key in data.keys())
    
    # Validate each tab separately - one function for each navigation page
    staff_errors, staff_data = _validate_staff_details(data, unique_id)
    
    # Only validate single fields if arrays are not present
    dependent_errors = {}
    dependent_data = {}
    if not has_dependents_array:
        dependent_errors, dependent_data = _validate_dependent_details(data)
        # Check if at least one dependent is required (only if no array data)
        if not dependent_data.get('rel_relationship'):
            dependent_errors['relationship'] = 'At least one dependent is required.'
    else:
        # Check if arrays have at least one record (only on final save)
        if is_final_save:
            dep_count = sum(1 for key in data.keys() if key.startswith('dependents[') and key.endswith('][relationship]'))
            if dep_count == 0:
                dependent_errors['dependents'] = 'At least one dependent is required.'
    
    account_errors = {}
    account_data = {}
    if not has_accounts_array:
        account_errors, account_data = _validate_account_details(data)
        # Check if at least one account is required (only if no array data)
        if not account_data.get('bank_status'):
            account_errors['bank_status'] = 'At least one account is required.'
    else:
        # Check if arrays have at least one record (only on final save)
        if is_final_save:
            acc_count = sum(1 for key in data.keys() if key.startswith('accounts[') and key.endswith('][bank_status]'))
            if acc_count == 0:
                account_errors['accounts'] = 'At least one account is required.'
    
    qualification_errors = {}
    qualification_data = {}
    if not has_qualifications_array:
        qualification_errors, qualification_data = _validate_qualification_details(data)
        # Check if at least one qualification is required (only if no array data)
        if not qualification_data.get('education_type'):
            qualification_errors['education_type'] = 'At least one qualification is required.'
    else:
        # Check if arrays have at least one record (only on final save)
        if is_final_save:
            qual_count = sum(1 for key in data.keys() if key.startswith('qualifications[') and key.endswith('][education_type]'))
            if qual_count == 0:
                qualification_errors['qualifications'] = 'At least one qualification is required.'
    
    experience_errors = {}
    experience_data = {}
    if not has_experiences_array:
        experience_errors, experience_data = _validate_experience_details(data)
        # Check if at least one experience is required (only if no array data)
        if not experience_data.get('exp_company_name'):
            experience_errors['exp_company_name'] = 'At least one experience is required.'
    else:
        # Check if arrays have at least one record (only on final save)
        if is_final_save:
            exp_count = sum(1 for key in data.keys() if key.startswith('experiences[') and key.endswith('][exp_company_name]'))
            if exp_count == 0:
                experience_errors['experiences'] = 'At least one experience is required.'
    
    asset_errors = {}
    asset_data = {}
    if not has_assets_array:
        asset_errors, asset_data = _validate_asset_vehicle_details(data)
        # Check if at least one asset is required (only if no array data)
        if not asset_data.get('asset_name'):
            asset_errors['asset_name'] = 'At least one asset is required.'
    else:
        # Check if arrays have at least one record (only on final save)
        if is_final_save:
            asset_count = sum(1 for key in data.keys() if key.startswith('assets[') and key.endswith('][asset_name]'))
            if asset_count == 0:
                asset_errors['assets'] = 'At least one asset is required.'
    
    # Validate arrays if present
    if has_dependents_array:
        index = 0
        dep_count = 0
        while f'dependents[{index}][relationship]' in data:
            dep_errors, _ = _validate_dependent_details({
                'relationship': data.get(f'dependents[{index}][relationship]', ''),
                'rel_name': data.get(f'dependents[{index}][rel_name]', ''),
                'rel_gender': data.get(f'dependents[{index}][rel_gender]', ''),
                'rel_date_of_birth': data.get(f'dependents[{index}][rel_date_of_birth]', ''),
                'rel_aadhar_no': data.get(f'dependents[{index}][rel_aadhar_no]', ''),
                'occupation': data.get(f'dependents[{index}][occupation]', ''),
                'standard': data.get(f'dependents[{index}][standard]', ''),
                'school': data.get(f'dependents[{index}][school]', ''),
                'existing_illness': data.get(f'dependents[{index}][existing_illness]', ''),
                'description': data.get(f'dependents[{index}][description]', ''),
                'existing_insurance': data.get(f'dependents[{index}][existing_insurance]', ''),
                'insurance_no': data.get(f'dependents[{index}][insurance_no]', ''),
                'physically_challenged': data.get(f'dependents[{index}][physically_challenged]', ''),
                'remarks': data.get(f'dependents[{index}][remarks]', ''),
            })
            for key, value in dep_errors.items():
                dependent_errors[f'dependents[{index}][{key}]'] = value
            if not dep_errors:  # Only count if no errors
                dep_count += 1
            index += 1
        if dep_count == 0 and index > 0 and is_final_save:
            # If we found array keys but all had errors, show general message (only on final save)
            if not dependent_errors:
                dependent_errors['dependents'] = 'At least one valid dependent is required.'
    
    if has_accounts_array:
        index = 0
        while f'accounts[{index}][bank_status]' in data:
            acc_errors, _ = _validate_account_details({
                'bank_status': data.get(f'accounts[{index}][bank_status]', ''),
                'salary_type': data.get(f'accounts[{index}][salary_type]', ''),
                'accountant_name': data.get(f'accounts[{index}][accountant_name]', ''),
                'account_no': data.get(f'accounts[{index}][account_no]', ''),
                'bank_name': data.get(f'accounts[{index}][bank_name]', ''),
                'ifsc_code': data.get(f'accounts[{index}][ifsc_code]', ''),
                'bank_contact_no': data.get(f'accounts[{index}][bank_contact_no]', ''),
                'bank_address': data.get(f'accounts[{index}][bank_address]', ''),
            })
            for key, value in acc_errors.items():
                account_errors[f'accounts[{index}][{key}]'] = value
            index += 1
        if index == 0:
            account_errors['accounts'] = 'At least one account is required.'
    
    if has_qualifications_array:
        index = 0
        qual_count = 0
        while f'qualifications[{index}][education_type]' in data:
            qual_errors, _ = _validate_qualification_details({
                'education_type': data.get(f'qualifications[{index}][education_type]', ''),
                'degree': data.get(f'qualifications[{index}][degree]', ''),
                'college_name': data.get(f'qualifications[{index}][college_name]', ''),
                'year_passing': data.get(f'qualifications[{index}][year_passing]', ''),
                'percentage': data.get(f'qualifications[{index}][percentage]', ''),
                'university': data.get(f'qualifications[{index}][university]', ''),
            })
            for key, value in qual_errors.items():
                qualification_errors[f'qualifications[{index}][{key}]'] = value
            if not qual_errors:  # Only count if no errors
                qual_count += 1
            index += 1
        if qual_count == 0 and index > 0 and is_final_save:
            if not qualification_errors:
                qualification_errors['qualifications'] = 'At least one valid qualification is required.'
    
    if has_experiences_array:
        index = 0
        exp_count = 0
        while f'experiences[{index}][exp_company_name]' in data:
            exp_errors, _ = _validate_experience_details({
                'exp_company_name': data.get(f'experiences[{index}][exp_company_name]', ''),
                'designation_name': data.get(f'experiences[{index}][designation_name]', ''),
                'salary_amt': data.get(f'experiences[{index}][salary_amt]', ''),
                'join_month': data.get(f'experiences[{index}][join_month]', ''),
                'relieve_month': data.get(f'experiences[{index}][relieve_month]', ''),
                'exp': data.get(f'experiences[{index}][exp]', ''),
            })
            for key, value in exp_errors.items():
                experience_errors[f'experiences[{index}][{key}]'] = value
            if not exp_errors:  # Only count if no errors
                exp_count += 1
            index += 1
        if exp_count == 0 and index > 0 and is_final_save:
            if not experience_errors:
                experience_errors['experiences'] = 'At least one valid experience is required.'
    
    if has_assets_array:
        index = 0
        asset_count = 0
        while f'assets[{index}][asset_name]' in data:
            asset_errors_temp, _ = _validate_asset_vehicle_details({
                'asset_name': data.get(f'assets[{index}][asset_name]', ''),
                'item_no': data.get(f'assets[{index}][item_no]', ''),
                'qty': data.get(f'assets[{index}][qty]', ''),
                'status': data.get(f'assets[{index}][status]', ''),
                'veh_reg_no': data.get(f'assets[{index}][veh_reg_no]', ''),
                'license_mode': data.get(f'assets[{index}][license_mode]', ''),
                'dri_license_no': data.get(f'assets[{index}][dri_license_no]', ''),
                'valid_from': data.get(f'assets[{index}][valid_from]', ''),
                'valid_to': data.get(f'assets[{index}][valid_to]', ''),
                'vehicle_type': data.get(f'assets[{index}][vehicle_type]', ''),
                'vehicle_company': data.get(f'assets[{index}][vehicle_company]', ''),
                'vehicle_owner': data.get(f'assets[{index}][vehicle_owner]', ''),
                'registration_year': data.get(f'assets[{index}][registration_year]', ''),
                'rc_no': data.get(f'assets[{index}][rc_no]', ''),
                'rc_validity_from': data.get(f'assets[{index}][rc_validity_from]', ''),
                'rc_validity_to': data.get(f'assets[{index}][rc_validity_to]', ''),
                'ins_no': data.get(f'assets[{index}][ins_no]', ''),
                'ins_validity_from': data.get(f'assets[{index}][ins_validity_from]', ''),
                'ins_validity_to': data.get(f'assets[{index}][ins_validity_to]', ''),
            })
            for key, value in asset_errors_temp.items():
                asset_errors[f'assets[{index}][{key}]'] = value
            if not asset_errors_temp:  # Only count if no errors
                asset_count += 1
            index += 1
        if asset_count == 0 and index > 0 and is_final_save:
            if not asset_errors:
                asset_errors['assets'] = 'At least one valid asset is required.'
    
    # Combine all errors
    all_errors = {}
    all_errors.update(staff_errors)
    all_errors.update(dependent_errors)
    all_errors.update(account_errors)
    all_errors.update(qualification_errors)
    all_errors.update(experience_errors)
    all_errors.update(asset_errors)
    
    # Combine all validated data (only for single records, arrays handled separately)
    validated_data = {}
    validated_data.update(staff_data)
    if not has_dependents_array:
        validated_data.update(dependent_data)
    if not has_accounts_array:
        validated_data.update(account_data)
    if not has_qualifications_array:
        validated_data.update(qualification_data)
    if not has_experiences_array:
        validated_data.update(experience_data)
    if not has_assets_array:
        validated_data.update(asset_data)
    
    # Validate required file uploads (only on final save)
    if is_final_save:
        profile_image = request.FILES.get('profile_image')
        if not profile_image:
            all_errors['profile_image'] = 'Profile image is required.'
        
        # Check file uploads for arrays
        if has_qualifications_array:
            index = 0
            has_qual_files = False
            while f'qualifications[{index}][education_type]' in data:
                qual_file = request.FILES.get(f'qualifications[{index}][qualification_docs]')
                if qual_file:
                    has_qual_files = True
                    break
                index += 1
            if not has_qual_files:
                qualification_errors['qualification_docs'] = 'At least one qualification document is required.'
        else:
            qualification_docs_list = request.FILES.getlist('qualification_docs')
            if not qualification_docs_list or len(qualification_docs_list) == 0:
                all_errors['qualification_docs'] = 'Document upload is required.'
        
        if has_experiences_array:
            index = 0
            has_exp_files = False
            while f'experiences[{index}][exp_company_name]' in data:
                exp_file = request.FILES.get(f'experiences[{index}][experience_docs]')
                if exp_file:
                    has_exp_files = True
                    break
                index += 1
            if not has_exp_files:
                experience_errors['experience_docs'] = 'At least one experience document is required.'
        else:
            experience_docs_list = request.FILES.getlist('experience_docs')
            if not experience_docs_list or len(experience_docs_list) == 0:
                all_errors['experience_docs'] = 'Document upload is required.'
    
    # Update all_errors with array errors
    all_errors.update(qualification_errors)
    all_errors.update(experience_errors)
    
    # If there are validation errors, return them
    if all_errors:
        return JsonResponse({'status': 0, 'errors': all_errors, 'msg': 'Validation failed. Please check all fields.'}, status=400)
    
    # Create or update Employee
    try:
        employee, created = Employee.objects.update_or_create(
            unique_id=unique_id,
            defaults={
                'staff_name': validated_data['staff_name'],
                'staff_id': validated_data['staff_id'],
                'gender': validated_data['gender'],
                'father_name': validated_data['father_name'],
                'date_of_birth': validated_data['date_of_birth'],
                'document_date_of_birth': validated_data['document_date_of_birth'],
                'age': validated_data['age'],
                'marital_status': validated_data['marital_status'],
                'personal_contact': validated_data['personal_contact'],
                'office_contact': validated_data['office_contact'],
                'personal_email': validated_data['personal_email'],
                'office_email': validated_data['office_email'],
                'aadhar_no': validated_data['aadhar_no'],
                'pan_no': validated_data['pan_no'],
                'medical_claim': validated_data['medical_claim'],
                'blood_group': validated_data['blood_group'],
                'qualification': validated_data['qualification'],
                'present_country': validated_data['present_country'],
                'present_state': validated_data['present_state'],
                'present_city': validated_data['present_city'],
                'present_building': validated_data['present_building'],
                'present_street': validated_data['present_street'],
                'present_area': validated_data['present_area'],
                'present_pincode': validated_data['present_pincode'],
                'permanent_country': validated_data['permanent_country'],
                'permanent_state': validated_data['permanent_state'],
                'permanent_city': validated_data['permanent_city'],
                'permanent_building': validated_data['permanent_building'],
                'permanent_street': validated_data['permanent_street'],
                'permanent_area': validated_data['permanent_area'],
                'permanent_pincode': validated_data['permanent_pincode'],
                'date_of_join': validated_data['date_of_join'],
                'designation': validated_data['designation'],
                'department': validated_data['department'],
                'work_location': validated_data['work_location'],
                'esi_no': validated_data['esi_no'],
                'pf_no': validated_data['pf_no'],
                'biometric_id': validated_data['biometric_id'],
                'company': validated_data['company'],
                'salary_category': validated_data['salary_category'],
                'premises_type': validated_data['premises_type'],
                'branch': validated_data['branch'],
                'attendance_setting': validated_data['attendance_setting'],
                'reporting_officer': validated_data['reporting_officer'],
            }
        )
        
        # Profile image is required, so it should always be present after validation
        profile_image = request.FILES.get('profile_image')
        if profile_image:
            employee.profile_image = profile_image
            employee.save()
        
        # ========== DEPENDENT DETAILS ==========
        # Delete existing dependents for this employee (if editing)
        EmployeeDependent.objects.filter(employee=employee).delete()
        
        # Save multiple dependents
        index = 0
        while f'dependents[{index}][relationship]' in data:
            dep_data = {
                'relationship': data.get(f'dependents[{index}][relationship]', '').strip(),
                'name': data.get(f'dependents[{index}][rel_name]', '').strip(),
                'gender': data.get(f'dependents[{index}][rel_gender]', '').strip(),
                'date_of_birth': _parse_date(data.get(f'dependents[{index}][rel_date_of_birth]', '').strip()),
                'aadhar_no': data.get(f'dependents[{index}][rel_aadhar_no]', '').strip().replace(' ', ''),
                'occupation': data.get(f'dependents[{index}][occupation]', '').strip(),
                'standard': data.get(f'dependents[{index}][standard]', '').strip(),
                'school': data.get(f'dependents[{index}][school]', '').strip(),
                'existing_illness': data.get(f'dependents[{index}][existing_illness]', '').strip(),
                'description': data.get(f'dependents[{index}][description]', '').strip(),
                'existing_insurance': data.get(f'dependents[{index}][existing_insurance]', '').strip(),
                'insurance_no': data.get(f'dependents[{index}][insurance_no]', '').strip(),
                'physically_challenged': data.get(f'dependents[{index}][physically_challenged]', '').strip(),
                'remarks': data.get(f'dependents[{index}][remarks]', '').strip(),
            }
            if dep_data['relationship']:  # Only save if relationship is provided
                EmployeeDependent.objects.create(employee=employee, **dep_data)
            index += 1
        
        # ========== ACCOUNT DETAILS ==========
        # Delete existing accounts for this employee (if editing)
        EmployeeAccountInfo.objects.filter(employee=employee).delete()
        
        # Save multiple accounts (though typically one per employee)
        index = 0
        while f'accounts[{index}][bank_status]' in data:
            acc_data = {
                'bank_status': data.get(f'accounts[{index}][bank_status]', '').strip(),
                'salary_type': data.get(f'accounts[{index}][salary_type]', '').strip(),
                'accountant_name': data.get(f'accounts[{index}][accountant_name]', '').strip(),
                'account_no': data.get(f'accounts[{index}][account_no]', '').strip(),
                'bank_name': data.get(f'accounts[{index}][bank_name]', '').strip(),
                'ifsc_code': data.get(f'accounts[{index}][ifsc_code]', '').strip().upper(),
                'contact_no': data.get(f'accounts[{index}][bank_contact_no]', '').strip(),
                'bank_address': data.get(f'accounts[{index}][bank_address]', '').strip(),
            }
            if acc_data['bank_status']:  # Only save if bank_status is provided
                EmployeeAccountInfo.objects.create(employee=employee, **acc_data)
            index += 1
        
        # ========== QUALIFICATION DETAILS ==========
        # Delete existing qualifications for this employee (if editing)
        EmployeeQualification.objects.filter(employee=employee).delete()
        
        # Save multiple qualifications
        index = 0
        while f'qualifications[{index}][education_type]' in data:
            qual_data = {
                'education_type': data.get(f'qualifications[{index}][education_type]', '').strip(),
                'degree': data.get(f'qualifications[{index}][degree]', '').strip(),
                'college_name': data.get(f'qualifications[{index}][college_name]', '').strip(),
                'year_of_passing': data.get(f'qualifications[{index}][year_passing]', '').strip(),
                'percentage': data.get(f'qualifications[{index}][percentage]', '').strip(),
                'university': data.get(f'qualifications[{index}][university]', '').strip(),
            }
            # Handle file upload
            qual_file = request.FILES.get(f'qualifications[{index}][qualification_docs]')
            if qual_data['education_type']:  # Only save if education_type is provided
                EmployeeQualification.objects.create(
                    employee=employee,
                    documents=qual_file,
                    **qual_data
                )
            index += 1
        
        # ========== EXPERIENCE DETAILS ==========
        # Delete existing experiences for this employee (if editing)
        EmployeeExperience.objects.filter(employee=employee).delete()
        
        # Save multiple experiences
        index = 0
        while f'experiences[{index}][exp_company_name]' in data:
            exp_data = {
                'company_name': data.get(f'experiences[{index}][exp_company_name]', '').strip(),
                'designation': data.get(f'experiences[{index}][designation_name]', '').strip(),
                'salary': data.get(f'experiences[{index}][salary_amt]', '').strip(),
                'joining_month': data.get(f'experiences[{index}][join_month]', '').strip(),
                'relieving_month': data.get(f'experiences[{index}][relieve_month]', '').strip(),
                'experience_years': data.get(f'experiences[{index}][exp]', '').strip(),
            }
            # Handle file upload
            exp_file = request.FILES.get(f'experiences[{index}][experience_docs]')
            if exp_data['company_name']:  # Only save if company_name is provided
                EmployeeExperience.objects.create(
                    employee=employee,
                    documents=exp_file,
                    **exp_data
                )
            index += 1
        
        # ========== ASSET/VEHICLE DETAILS ==========
        # Delete existing assets for this employee (if editing)
        EmployeeAssetAssignment.objects.filter(employee=employee).delete()
        
        # Save multiple assets
        index = 0
        while f'assets[{index}][asset_name]' in data:
            asset_data = {
                'asset_name': data.get(f'assets[{index}][asset_name]', '').strip(),
                'serial_no': data.get(f'assets[{index}][item_no]', '').strip(),
                'quantity': int(data.get(f'assets[{index}][qty]', '1').strip() or '1'),
                'status': data.get(f'assets[{index}][status]', EmployeeAssetAssignment.STATUS_ISSUED).strip(),
                'vehicle_reg_no': data.get(f'assets[{index}][veh_reg_no]', '').strip(),
                'license_mode': data.get(f'assets[{index}][license_mode]', '').strip(),
                'license_no': data.get(f'assets[{index}][dri_license_no]', '').strip(),
                'license_valid_from': _parse_date(data.get(f'assets[{index}][valid_from]', '').strip()),
                'license_valid_to': _parse_date(data.get(f'assets[{index}][valid_to]', '').strip()),
            }
            if asset_data['asset_name']:  # Only save if asset_name is provided
                EmployeeAssetAssignment.objects.create(employee=employee, **asset_data)
            index += 1
        
        # Vehicle Details (OneToOne relationship) - Get from last asset or validated_data
        vehicle_data = {}
        if has_assets_array:
            # Get vehicle details from last asset in array
            index = 0
            last_index = -1
            while f'assets[{index}][asset_name]' in data:
                last_index = index
                index += 1
            if last_index >= 0:
                vehicle_data = {
                    'vehicle_type': data.get(f'assets[{last_index}][vehicle_type]', '').strip(),
                    'vehicle_company': data.get(f'assets[{last_index}][vehicle_company]', '').strip(),
                    'vehicle_owner': data.get(f'assets[{last_index}][vehicle_owner]', '').strip(),
                    'registration_year': _parse_date(data.get(f'assets[{last_index}][registration_year]', '').strip()),
                    'rc_no': data.get(f'assets[{last_index}][rc_no]', '').strip(),
                    'rc_validity_from': _parse_date(data.get(f'assets[{last_index}][rc_validity_from]', '').strip()),
                    'rc_validity_to': _parse_date(data.get(f'assets[{last_index}][rc_validity_to]', '').strip()),
                    'insurance_no': data.get(f'assets[{last_index}][ins_no]', '').strip(),
                    'insurance_validity_from': _parse_date(data.get(f'assets[{last_index}][ins_validity_from]', '').strip()),
                    'insurance_validity_to': _parse_date(data.get(f'assets[{last_index}][ins_validity_to]', '').strip()),
                }
        else:
            # Get from validated_data (single record)
            vehicle_data = {
                'vehicle_type': validated_data.get('vehicle_type', ''),
                'vehicle_company': validated_data.get('vehicle_company', ''),
                'vehicle_owner': validated_data.get('vehicle_owner', ''),
                'registration_year': validated_data.get('registration_year'),
                'rc_no': validated_data.get('rc_no', ''),
                'rc_validity_from': validated_data.get('rc_validity_from'),
                'rc_validity_to': validated_data.get('rc_validity_to'),
                'insurance_no': validated_data.get('ins_no', ''),
                'insurance_validity_from': validated_data.get('ins_validity_from'),
                'insurance_validity_to': validated_data.get('ins_validity_to'),
            }
        
        # Only save vehicle details if at least vehicle_type is provided
        if vehicle_data.get('vehicle_type'):
            EmployeeVehicleDetail.objects.update_or_create(
                employee=employee,
                defaults=vehicle_data
            )
        
        action = 'created' if created else 'updated'
        return JsonResponse({
            'status': 1,
            'msg': f'Employee {action} successfully.',
            'unique_id': employee.unique_id,
            'employee_id': employee.id,
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 0,
            'msg': 'An error occurred while saving employee data.',
            'error': str(e)
        }, status=500)
